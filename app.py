import os
import uuid
import math
import pandas as pd
from io import BytesIO
from datetime import datetime
import requests as http_req
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, make_response
)
from models import (
    db, Customer, VehicleRate, VehicleCapacity, SurchargeRule,
    JointDeliveryRate, SystemConfig, ProductMaster, StoreMaster,
    ShippingHistory, CalculationResult, OurCenter,
    TransferRate, HubVehicleRate, SynergyRoute, CustomerStorageCenter,
    DestinationCoord, DeliveryZoneMapping, VehicleDistanceRate
)
from calculator import (
    calculate_from_history, summarize_results,
    extract_sido_sigungu, make_destination_key, normalize_sido,
    compute_joint_breakdown_detail,
    get_direct_plt_threshold,
)

# 시도명 정규화 테이블: 단축형/약식명 → 표준 단축형 (양방향 정규화용)
_SIDO_NORM = {
    '서울특별시': '서울', '서울시': '서울', '서울': '서울',
    '부산광역시': '부산', '부산': '부산',
    '대구광역시': '대구', '대구': '대구',
    '인천광역시': '인천', '인천시': '인천', '인천': '인천',
    '광주광역시': '광주', '광주': '광주',
    '대전광역시': '대전', '대전': '대전',
    '울산광역시': '울산', '울산': '울산',
    '세종특별자치시': '세종', '세종시': '세종', '세종': '세종',
    '경기도': '경기', '경기': '경기',
    '강원특별자치도': '강원', '강원도': '강원', '강원': '강원',
    '충청북도': '충북', '충북': '충북',
    '충청남도': '충남', '충남': '충남',
    '전라북도': '전북', '전북': '전북', '전북특별자치도': '전북',
    '전라남도': '전남', '전남': '전남',
    '경상북도': '경북', '경북': '경북',
    '경상남도': '경남', '경남': '경남',
    '제주특별자치도': '제주', '제주도': '제주', '제주': '제주',
}

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///delivery_pricing.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'hanex-delivery-2024'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
db.init_app(app)

with app.app_context():
    db.create_all()
    with db.engine.connect() as _conn:
        for _sql in [
            "ALTER TABLE our_center ADD COLUMN is_main_center BOOLEAN DEFAULT 0",
            "ALTER TABLE calculation_results ADD COLUMN transfer_cost INTEGER",
            "ALTER TABLE calculation_results ADD COLUMN hub_cost INTEGER",
        ]:
            try:
                _conn.execute(db.text(_sql))
                _conn.commit()
            except Exception:
                pass

# 기본 차량 적재량 (없으면 자동 삽입)
DEFAULT_CAPACITIES = [
    ('11톤',   18, 0),
    ('5톤장축', 12, 1),
    ('5톤',    10, 2),
    ('3.5톤',   4, 3),
    ('2.5톤',   3, 4),
    ('1.4톤',   2, 5),
    ('1톤',     1, 6),
    ('퀵',    0.5, 7),
]

DEFAULT_CONFIGS = [
    ('direct_plt_threshold', '3', '직송 전환 기준 PLT 수 (이 값 이상이면 직송)'),
    ('kakao_api_key', 'd3814275e891fd92e98e84d5d80252d6', '카카오 주소→좌표 변환 REST API 키'),
    ('stops_per_vehicle', '8', '공동배송 차량당 평균 도착지(점포) 수'),
]

with app.app_context():
    if VehicleCapacity.query.count() == 0:
        for vt, mp, so in DEFAULT_CAPACITIES:
            db.session.add(VehicleCapacity(vehicle_type=vt, max_plt=mp, sort_order=so))
    for key, val, desc in DEFAULT_CONFIGS:
        if not SystemConfig.query.filter_by(key=key).first():
            db.session.add(SystemConfig(key=key, value=val, description=desc))
    db.session.commit()


# ─── 화주 현황 ─────────────────────────────────────────────────────────────────

@app.route('/')
def analytics():
    from sqlalchemy import func as sqlfunc
    from calculator import compute_joint_breakdown_live

    customers = Customer.query.order_by(Customer.name).all()
    cust_map  = {c.id: c.name for c in customers}

    # 공통 설정
    _spv_cfg = SystemConfig.query.filter_by(key='stops_per_vehicle').first()
    stops_per_vehicle = int(_spv_cfg.value) if _spv_cfg else 8
    _main_ctr = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).first()
    _main_code = _main_ctr.center_code if _main_ctr else None

    # 고객별 최신 batch_id 조회 (중복 배치 방지)
    _batch_rows = db.session.query(
        CalculationResult.customer_id,
        CalculationResult.batch_id,
        sqlfunc.max(CalculationResult.calc_date).label('max_date'),
    ).filter(CalculationResult.cost_per_box.isnot(None)).group_by(
        CalculationResult.customer_id, CalculationResult.batch_id
    ).order_by(CalculationResult.customer_id, sqlfunc.max(CalculationResult.calc_date).desc()).all()

    cust_batch_map = {}
    for _br in _batch_rows:
        if _br.customer_id not in cust_batch_map:
            cust_batch_map[_br.customer_id] = _br.batch_id

    latest_batch_ids = list(cust_batch_map.values()) if cust_batch_map else [None]

    rows = db.session.query(
        CalculationResult.customer_id,
        sqlfunc.count(CalculationResult.id).label('cnt'),
        sqlfunc.sum(CalculationResult.total_box_qty).label('boxes'),
        sqlfunc.sum(CalculationResult.total_plt_decimal).label('plt'),
        sqlfunc.count(sqlfunc.distinct(CalculationResult.shipping_date)).label('days'),
        sqlfunc.sum(db.case((CalculationResult.delivery_mode == '직송', 1), else_=0)).label('direct_cnt'),
        sqlfunc.sum(db.case((CalculationResult.delivery_mode == '직송',
                              CalculationResult.delivery_cost), else_=0)).label('direct_cost'),
        sqlfunc.max(CalculationResult.calc_date).label('last_calc'),
    ).filter(
        CalculationResult.cost_per_box.isnot(None),
        CalculationResult.batch_id.in_(latest_batch_ids),
    ).group_by(CalculationResult.customer_id).all()

    customer_stats = []
    for r in rows:
        cid         = r.customer_id
        boxes       = int(r.boxes or 0)
        days        = int(r.days or 1)
        direct_cost = int(r.direct_cost or 0)
        _bid        = cust_batch_map.get(cid)

        joint_cnt = r.cnt - int(r.direct_cnt or 0)

        if joint_cnt > 0:
            _jd_f = [
                CalculationResult.customer_id == cid,
                CalculationResult.delivery_mode == '공동배송',
                CalculationResult.shipping_date.isnot(None),
            ]
            if _bid:
                _jd_f.append(CalculationResult.batch_id == _bid)
            joint_days = db.session.query(
                sqlfunc.count(sqlfunc.distinct(CalculationResult.shipping_date))
            ).filter(*_jd_f).scalar() or 1

            transfer_cost, _ = compute_joint_breakdown_live(cid, _main_code, stops_per_vehicle, db.session, batch_id=_bid)
            hub_vehicle_cost  = _hub_vehicle_daily_cost(cid, stops_per_vehicle) * joint_days
        else:
            transfer_cost    = 0
            hub_vehicle_cost = 0

        total_cost = direct_cost + transfer_cost + hub_vehicle_cost

        customer_stats.append({
            'id':              cid,
            'name':            cust_map.get(cid, f'고객#{cid}'),
            'cnt':             r.cnt,
            'boxes':           boxes,
            'total_cost':      total_cost,
            'avg_cpb':         round(total_cost / max(1, boxes), 1),
            'direct_pct':      round(int(r.direct_cnt or 0) / r.cnt * 100) if r.cnt else 0,
            'last_calc':       r.last_calc,
            'daily_avg_boxes': round(boxes / max(1, days)),
            'daily_avg_plt':   round(float(r.plt or 0) / max(1, days), 1),
        })

    customer_stats.sort(key=lambda x: x['name'])

    return render_template('analytics.html',
        customers=customers,
        customer_stats=customer_stats,
    )


def _hub_vehicle_daily_cost(customer_id, stops_per_vehicle, threshold=None):
    """
    지도와 완전히 동일한 grandTotal 산출.
    threshold 지정 시 해당 기준으로 공동배송 분류 후 루트 재계산.
    """
    hub_centers = _customer_map_data(customer_id, _return_raw=True, threshold=threshold)
    if not hub_centers:
        return 0

    dist_rate_map = {}
    for dr in VehicleDistanceRate.query.all():
        dist_rate_map.setdefault(dr.vehicle_type, {})[dr.km] = dr.unit_price
    if not dist_rate_map:
        return 0

    _HUB_CAP = [(2, '1T'), (3, '2.5T'), (5, '3.5T'), (10, '5T'), (16, '11T')]
    total_hub = 0

    for hub in hub_centers:
        hub_lat = float(hub['lat'])
        hub_lon = float(hub['lon'])
        zones_raw = [
            (float(z['lat']), float(z['lon']), float(z.get('avg_plt') or 0))
            for z in hub['zones']
            if z.get('lat') and z.get('lon')
        ]
        if not zones_raw:
            continue

        sorted_z = sorted(zones_raw, key=lambda z: math.atan2(z[1] - hub_lon, z[0] - hub_lat))

        for v in range(math.ceil(len(sorted_z) / stops_per_vehicle)):
            batch = sorted_z[v * stops_per_vehicle:(v + 1) * stops_per_vehicle]

            unvisited = list(batch)
            route = []
            cur_lat, cur_lon = hub_lat, hub_lon
            while unvisited:
                bi, bd = 0, float('inf')
                for i, (lat, lon, _) in enumerate(unvisited):
                    dx = math.radians(lat - cur_lat)
                    dy = math.radians(lon - cur_lon)
                    a = (math.sin(dx / 2) ** 2
                         + math.cos(math.radians(cur_lat)) * math.cos(math.radians(lat))
                         * math.sin(dy / 2) ** 2)
                    d = 6371 * 2 * math.asin(math.sqrt(a))
                    if d < bd:
                        bd, bi = d, i
                stop = unvisited.pop(bi)
                route.append(stop)
                cur_lat, cur_lon = stop[0], stop[1]

            km = 0.0
            cur_lat, cur_lon = hub_lat, hub_lon
            for lat, lon, _ in route:
                dx = math.radians(lat - cur_lat)
                dy = math.radians(lon - cur_lon)
                a = (math.sin(dx / 2) ** 2
                     + math.cos(math.radians(cur_lat)) * math.cos(math.radians(lat))
                     * math.sin(dy / 2) ** 2)
                km += 6371 * 2 * math.asin(math.sqrt(a))
                cur_lat, cur_lon = lat, lon
            road_km = max(1, min(1000, round(km * 1.3)))

            route_plt = sum(ap for _, _, ap in route)
            if route_plt <= 0:
                continue

            sel_vt, sel_cap = '11T', 16
            for max_p, vt in _HUB_CAP:
                if max_p >= route_plt:
                    sel_vt, sel_cap = vt, max_p
                    break

            up = (dist_rate_map.get(sel_vt) or {}).get(road_km)
            if up:
                total_hub += round(up * (route_plt / sel_cap))

    return total_hub


@app.route('/help')
def help_page():
    import os
    ss_dir = os.path.join(app.static_folder, 'help', 'screenshots')
    taken = set()
    if os.path.isdir(ss_dir):
        taken = {f.rsplit('.', 1)[0] for f in os.listdir(ss_dir) if f.endswith('.png')}
    return render_template('help.html', screenshots={k: True for k in taken})


@app.route('/analytics/<int:customer_id>')
def analytics_detail(customer_id):
    from sqlalchemy import func as sqlfunc
    import json

    customer = Customer.query.get_or_404(customer_id)

    # 최신 배치만 사용 (여러 배치가 쌓이면 중복 합산 방지)
    latest_batch = db.session.query(CalculationResult.batch_id).filter(
        CalculationResult.customer_id == customer_id
    ).order_by(CalculationResult.calc_date.desc()).first()
    latest_batch_id = latest_batch[0] if latest_batch else None

    f = (CalculationResult.customer_id == customer_id)
    if latest_batch_id:
        f = f & (CalculationResult.batch_id == latest_batch_id)

    total_results = CalculationResult.query.filter(f).count()

    # overall_avg_cpb는 total_cost / total_boxes 이후에 계산 (아래로 이동)
    overall_avg_cpb = 0

    mode_rows = db.session.query(
        CalculationResult.delivery_mode,
        sqlfunc.count(CalculationResult.id).label('cnt'),
        sqlfunc.sum(CalculationResult.total_box_qty).label('boxes'),
        sqlfunc.sum(CalculationResult.total_plt_decimal).label('plt'),
    ).filter(f).group_by(CalculationResult.delivery_mode).all()
    direct_cnt   = next((r.cnt for r in mode_rows if r.delivery_mode == '직송'), 0)
    joint_cnt    = next((r.cnt for r in mode_rows if r.delivery_mode == '공동배송'), 0)
    direct_boxes = int(next((r.boxes for r in mode_rows if r.delivery_mode == '직송'), 0) or 0)
    joint_boxes  = int(next((r.boxes for r in mode_rows if r.delivery_mode == '공동배송'), 0) or 0)
    direct_plt   = round(float(next((r.plt for r in mode_rows if r.delivery_mode == '직송'), 0) or 0), 1)
    joint_plt    = round(float(next((r.plt for r in mode_rows if r.delivery_mode == '공동배송'), 0) or 0), 1)
    direct_pct   = round(direct_cnt / (direct_cnt + joint_cnt) * 100) if (direct_cnt + joint_cnt) else 0

    total_boxes = int(db.session.query(
        sqlfunc.sum(CalculationResult.total_box_qty)
    ).filter(f).scalar() or 0)

    total_plt = round(float(db.session.query(
        sqlfunc.sum(CalculationResult.total_plt_decimal)
    ).filter(f).scalar() or 0), 1)

    direct_cost = int(db.session.query(sqlfunc.sum(CalculationResult.delivery_cost)).filter(f, CalculationResult.delivery_mode == '직송').scalar() or 0)

    # 현재 stops_per_vehicle로 이고비/변동용차비 실시간 계산
    _spv_cfg = SystemConfig.query.filter_by(key='stops_per_vehicle').first()
    stops_per_vehicle = int(_spv_cfg.value) if _spv_cfg else 8
    _main_ctr = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).first()
    _main_code = _main_ctr.center_code if _main_ctr else None

    # 공동배송 영업일수 & 변동용차비 일평균 (지도와 동일) → 총합 = 일평균 × 배송일수
    joint_days_cnt = db.session.query(
        sqlfunc.count(sqlfunc.distinct(CalculationResult.shipping_date))
    ).filter(f, CalculationResult.delivery_mode == '공동배송',
             CalculationResult.shipping_date.isnot(None)).scalar() or 1
    daily_avg_hub_cost = _hub_vehicle_daily_cost(customer_id, stops_per_vehicle) if joint_cnt > 0 else 0

    if joint_cnt > 0:
        _detail = compute_joint_breakdown_detail(customer_id, _main_code, stops_per_vehicle, db.session, batch_id=latest_batch_id)
        transfer_cost    = sum(item['total_cost'] for item in _detail['transfer'])
        hub_vehicle_cost = daily_avg_hub_cost * joint_days_cnt
    else:
        transfer_cost    = int(db.session.query(sqlfunc.sum(CalculationResult.transfer_cost)).filter(f, CalculationResult.delivery_mode == '공동배송').scalar() or 0)
        hub_vehicle_cost = 0

    # 공동배송 물류비 = 이고비 + 변동용차비 (detail 계산, 모달 총합과 일치)
    joint_cost  = transfer_cost + hub_vehicle_cost
    total_cost  = direct_cost + joint_cost
    # 박스당 평균단가 = 총물류비 ÷ 총박스수 (박스 수량 가중 평균)
    overall_avg_cpb = round(total_cost / total_boxes, 1) if total_boxes > 0 else 0

    daily_rows = db.session.query(
        CalculationResult.shipping_date,
        sqlfunc.sum(CalculationResult.total_box_qty).label('boxes'),
        sqlfunc.sum(CalculationResult.total_plt_decimal).label('plt'),
    ).filter(f, CalculationResult.shipping_date.isnot(None)).group_by(
        CalculationResult.shipping_date
    ).order_by(CalculationResult.shipping_date).all()

    weekday_rows = db.session.query(
        db.func.strftime('%w', CalculationResult.shipping_date).label('wd'),
        sqlfunc.sum(CalculationResult.total_box_qty).label('boxes'),
        sqlfunc.sum(CalculationResult.total_plt_decimal).label('plt'),
    ).filter(f, CalculationResult.shipping_date.isnot(None)).group_by('wd').all()
    wd_boxes = [0] * 7
    wd_plt   = [0.0] * 7
    for r in weekday_rows:
        idx = int(r.wd)
        wd_boxes[idx] = round(float(r.boxes or 0))
        wd_plt[idx]   = round(float(r.plt or 0), 1)

    # 요일별 출현 날짜 수 (하루 평균 산출용)
    wd_daycount_rows = db.session.query(
        db.func.strftime('%w', CalculationResult.shipping_date).label('wd'),
        sqlfunc.count(db.func.distinct(CalculationResult.shipping_date)).label('day_cnt'),
    ).filter(f, CalculationResult.shipping_date.isnot(None)).group_by('wd').all()
    wd_count = [1] * 7
    for r in wd_daycount_rows:
        wd_count[int(r.wd)] = max(1, int(r.day_cnt or 1))

    monthly_rows = db.session.query(
        db.func.strftime('%Y-%m', CalculationResult.shipping_date).label('ym'),
        sqlfunc.sum(CalculationResult.total_box_qty).label('boxes'),
        sqlfunc.sum(CalculationResult.total_plt_decimal).label('plt'),
    ).filter(f, CalculationResult.shipping_date.isnot(None)).group_by('ym').order_by('ym').all()

    VT_ORDER = ['11톤', '5톤장축', '5톤', '3.5톤', '2.5톤', '1.4톤', '1톤', '퀵']
    veh_raw = db.session.query(
        CalculationResult.vehicle_type,
        sqlfunc.count(CalculationResult.id).label('cnt')
    ).filter(f, CalculationResult.vehicle_type.isnot(None)).group_by(
        CalculationResult.vehicle_type
    ).all()
    veh_dict   = {r.vehicle_type: r.cnt for r in veh_raw}
    veh_sorted = [(vt, veh_dict[vt]) for vt in VT_ORDER if vt in veh_dict]
    veh_sorted += [(vt, cnt) for vt, cnt in veh_dict.items() if vt not in VT_ORDER]

    total_distinct_days  = len(daily_rows)

    # 직송 실제 배송 날짜 수 (일평균 분모)
    direct_days_cnt = db.session.query(
        sqlfunc.count(sqlfunc.distinct(CalculationResult.shipping_date))
    ).filter(f, CalculationResult.delivery_mode == '직송',
             CalculationResult.shipping_date.isnot(None)).scalar() or 1

    daily_avg_boxes       = round(total_boxes / total_distinct_days) if total_distinct_days > 0 else 0
    daily_avg_plt         = round(total_plt   / total_distinct_days, 1) if total_distinct_days > 0 else 0
    daily_avg_total_cost  = round(total_cost  / total_distinct_days) if total_distinct_days > 0 else 0
    daily_avg_direct_cost = round(direct_cost / direct_days_cnt)
    daily_avg_joint_cost  = round(joint_cost  / joint_days_cnt)
    direct_avg_boxes      = round(direct_boxes / direct_days_cnt)
    direct_avg_plt        = round(direct_plt   / direct_days_cnt, 1)
    joint_avg_boxes       = round(joint_boxes  / joint_days_cnt)
    joint_avg_plt         = round(joint_plt    / joint_days_cnt, 1)

    MON_FIRST = [1, 2, 3, 4, 5, 6, 0]
    chart_daily    = {
        'labels': [str(r.shipping_date) for r in daily_rows],
        'boxes':  [round(float(r.boxes or 0)) for r in daily_rows],
        'plt':    [round(float(r.plt or 0), 1) for r in daily_rows],
    }
    chart_weekday  = {
        'labels':    ['월', '화', '수', '목', '금', '토', '일'],
        'boxes':     [wd_boxes[i] for i in MON_FIRST],
        'plt':       [wd_plt[i]   for i in MON_FIRST],
        'avg_boxes': [round(wd_boxes[i] / wd_count[i]) for i in MON_FIRST],
        'avg_plt':   [round(wd_plt[i]   / wd_count[i], 1) for i in MON_FIRST],
    }
    chart_monthly  = {
        'labels': [r.ym for r in monthly_rows],
        'boxes':  [round(float(r.boxes or 0)) for r in monthly_rows],
        'plt':    [round(float(r.plt or 0), 1) for r in monthly_rows],
    }
    chart_mode = {
        'labels': ['직송', '공동배송'],
        'cnt':    [direct_cnt, joint_cnt],
        'boxes':  [direct_boxes, joint_boxes],
    }
    chart_veh = {
        'labels': [vt for vt, _ in veh_sorted],
        'data':   [cnt for _, cnt in veh_sorted],
    }

    # 시너지 요약 (자사 배송지시서 데이터가 있을 때만)
    synergy_summary = None
    if total_results > 0 and SynergyRoute.query.filter_by(car_flag=2).count() > 0:
        _sa = _run_synergy_analysis(customer_id)
        synergy_summary = {
            'overlap_pct':   _sa['overlap_pct'],
            'match_cnt':     _sa['match_cnt'],
            'total_cust':    _sa['total_cust'],
            'match_regions': _sa['match_regions'],
            'match_plt':     _sa['match_plt'],
        }

    return render_template('analytics_detail.html',
        customer=customer,
        total_results=total_results,
        total_boxes=total_boxes,
        total_plt=total_plt,
        daily_avg_boxes=daily_avg_boxes,
        daily_avg_plt=daily_avg_plt,
        overall_avg_cpb=overall_avg_cpb,
        total_cost=total_cost,
        direct_cost=direct_cost,
        joint_cost=joint_cost,
        transfer_cost=transfer_cost,
        hub_vehicle_cost=hub_vehicle_cost,
        stops_per_vehicle=stops_per_vehicle,
        direct_plt_threshold=get_direct_plt_threshold(db.session),
        total_distinct_days=total_distinct_days,
        daily_avg_total_cost=daily_avg_total_cost,
        daily_avg_direct_cost=daily_avg_direct_cost,
        daily_avg_hub_cost=daily_avg_hub_cost,
        daily_avg_joint_cost=daily_avg_joint_cost,
        direct_cnt=direct_cnt,
        joint_cnt=joint_cnt,
        direct_boxes=direct_boxes,
        joint_boxes=joint_boxes,
        direct_plt=direct_plt,
        joint_plt=joint_plt,
        direct_avg_boxes=direct_avg_boxes,
        direct_avg_plt=direct_avg_plt,
        joint_avg_boxes=joint_avg_boxes,
        joint_avg_plt=joint_avg_plt,
        direct_pct=direct_pct,
        synergy_summary=synergy_summary,
        chart_daily=json.dumps(chart_daily, ensure_ascii=False),
        chart_weekday=json.dumps(chart_weekday, ensure_ascii=False),
        chart_monthly=json.dumps(chart_monthly, ensure_ascii=False),
        chart_mode=json.dumps(chart_mode, ensure_ascii=False),
        chart_veh=json.dumps(chart_veh, ensure_ascii=False),
    )


@app.route('/analytics/<int:customer_id>/export')
def analytics_export(customer_id):
    """출고내역 전체를 재계산하여 행별 배송구분·단가를 엑셀로 반환."""
    from calculator import calculate_from_history
    customer = Customer.query.get_or_404(customer_id)

    # 메인 센터
    _spv_cfg = SystemConfig.query.filter_by(key='stops_per_vehicle').first()
    stops_per_vehicle = int(_spv_cfg.value) if _spv_cfg else 8
    _main_ctr = OurCenter.query.filter_by(is_main_center=True).first()
    main_code = _main_ctr.center_code if _main_ctr else None
    if not main_code:
        flash('메인 센터가 설정되어 있지 않습니다.', 'danger')
        return redirect(url_for('analytics_detail', customer_id=customer_id))

    history_rows = ShippingHistory.query.filter_by(customer_id=customer_id).all()
    if not history_rows:
        flash('출고내역이 없습니다.', 'danger')
        return redirect(url_for('analytics_detail', customer_id=customer_id))

    results, errors, error_rows = calculate_from_history(history_rows, customer_id, '', main_code, db.session)

    def _to_row(r):
        return {
            '납품일자':       r['shipping_date'],
            '점포코드':       r['store_code'] or '',
            '점포명':         r['store_name'] or '',
            '주소':           r['address'] or '',
            '도착지':         r['destination'] or '',
            '배송구분':       r['delivery_mode'] or '',
            '박스수':         r['total_box_qty'] or '',
            'PLT(소수)':      r['total_plt_decimal'] or '',
            'PLT(올림)':      r['total_plt_count'] or '',
            '차량종류':       r.get('vehicle_type') or '',
            '이고비(원)':     r.get('transfer_cost') or '',
            '변동용차비(원)': r.get('hub_cost') or '',
            '배송비합계(원)': r.get('delivery_cost') or '',
            '박스당단가(원)': r.get('cost_per_box') or '',
            '비고':           r.get('memo') or '',
        }

    # 정상 행 + 오류 행을 날짜순으로 합산
    all_rows = sorted(results + error_rows, key=lambda r: (str(r.get('shipping_date') or ''), r.get('store_name') or ''))
    df = pd.DataFrame([_to_row(r) for r in all_rows])

    # 오류 행 인덱스 (df 기준, header=1행)
    error_mode_set = {'오류'}
    error_excel_rows = {i + 2 for i, r in enumerate(all_rows) if r.get('delivery_mode') in error_mode_set}

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='배송단가산정')
        ws = writer.sheets['배송단가산정']

        # 오류 행 배경색 (연한 주황)
        from openpyxl.styles import PatternFill, Font
        err_fill = PatternFill(fill_type='solid', fgColor='FFE5CC')
        err_font = Font(color='CC4400')
        for row_idx in error_excel_rows:
            for cell in ws[row_idx]:
                cell.fill = err_fill
                cell.font = err_font

        # 열 너비 자동 조정
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    output.seek(0)

    filename = f'{customer.name}_배송단가산정_{datetime.now().strftime("%Y%m%d")}.xlsx'
    token = request.args.get('token', '')
    resp = make_response(send_file(output, download_name=filename, as_attachment=True,
                                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    if token:
        resp.set_cookie('export_done', token, max_age=30, samesite='Lax')
    return resp


@app.route('/analytics/<int:customer_id>/delete', methods=['POST'])
def analytics_delete(customer_id):
    if request.form.get('confirm_text') != '삭제':
        flash('삭제 확인 텍스트가 올바르지 않습니다.', 'danger')
        return redirect(url_for('analytics'))
    CalculationResult.query.filter_by(customer_id=customer_id).delete()
    db.session.commit()
    flash('산정 이력이 삭제되었습니다.', 'success')
    return redirect(url_for('analytics'))


# ─── 고객사 ────────────────────────────────────────────────────────────────────

@app.route('/customers')
def customer_list():
    customers = Customer.query.order_by(Customer.created_at.desc()).all()
    return render_template('customers/list.html', customers=customers)


@app.route('/customers/new', methods=['GET', 'POST'])
def customer_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        memo = request.form.get('memo', '').strip()
        if not name:
            flash('고객사명을 입력해주세요.', 'danger')
            return render_template('customers/form.html')
        c = Customer(name=name, memo=memo)
        db.session.add(c)
        db.session.commit()
        flash(f'고객사 [{name}]이(가) 등록되었습니다.', 'success')
        return redirect(url_for('customer_detail', cid=c.id))
    return render_template('customers/form.html')


@app.route('/customers/<int:cid>')
def customer_detail(cid):
    customer = Customer.query.get_or_404(cid)
    product_count = ProductMaster.query.filter_by(customer_id=cid).count()
    store_count = StoreMaster.query.filter_by(customer_id=cid).count()
    history_count = ShippingHistory.query.filter_by(customer_id=cid).count()
    result_count = CalculationResult.query.filter_by(customer_id=cid).count()
    return render_template('customers/detail.html',
                           customer=customer,
                           product_count=product_count,
                           store_count=store_count,
                           history_count=history_count,
                           result_count=result_count)


@app.route('/customers/<int:cid>/delete', methods=['POST'])
def customer_delete(cid):
    customer = Customer.query.get_or_404(cid)
    name = customer.name
    db.session.delete(customer)
    db.session.commit()
    flash(f'고객사 [{name}]이(가) 삭제되었습니다.', 'warning')
    return redirect(url_for('customer_list'))


# ─── 차량 단가 마스터 (센터별 매트릭스) ───────────────────────────────────────

@app.route('/masters/vehicle')
def vehicle_master():
    centers    = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).all()
    center_map = {c.center_code: c.center_name for c in centers}
    selected   = request.args.get('center', centers[0].center_code if centers else '')

    capacities   = VehicleCapacity.query.order_by(VehicleCapacity.sort_order).all()
    vehicle_types = [c.vehicle_type for c in capacities]

    # 선택된 센터의 매트릭스
    rates = VehicleRate.query.filter_by(center_code=selected).order_by(VehicleRate.destination).all()
    destinations = sorted(set(r.destination for r in rates))
    matrix = {}
    for r in rates:
        matrix.setdefault(r.destination, {})[r.vehicle_type] = r.unit_price

    # 센터별 도착지 수
    from sqlalchemy import func as sqlfunc
    stats_q = db.session.query(
        VehicleRate.center_code,
        sqlfunc.count(sqlfunc.distinct(VehicleRate.destination))
    ).group_by(VehicleRate.center_code).all()
    center_stats = {code: cnt for code, cnt in stats_q}

    return render_template('masters/vehicle.html',
                           centers=centers, center_map=center_map, selected=selected,
                           destinations=destinations, vehicle_types=vehicle_types,
                           matrix=matrix, capacities=capacities,
                           center_stats=center_stats)


@app.route('/masters/vehicle/save-matrix', methods=['POST'])
def vehicle_matrix_save():
    center_code = request.form.get('center_code', '').strip()
    dests  = request.form.getlist('dest')
    vts    = request.form.getlist('vt')
    prices = request.form.getlist('price')
    if not center_code:
        flash('센터를 선택해주세요.', 'danger')
        return redirect(url_for('vehicle_master'))
    saved = deleted = 0
    for dest, vt, p_str in zip(dests, vts, prices):
        p_str = p_str.replace(',', '').strip()
        existing = VehicleRate.query.filter_by(
            center_code=center_code, destination=dest, vehicle_type=vt
        ).first()
        if p_str and int(p_str) > 0:
            price = int(p_str)
            if existing:
                existing.unit_price = price
            else:
                db.session.add(VehicleRate(
                    center_code=center_code, destination=dest,
                    vehicle_type=vt, unit_price=price
                ))
            saved += 1
        elif existing:
            db.session.delete(existing)
            deleted += 1
    db.session.commit()
    flash(f'직송 단가 저장 — {saved}건 저장, {deleted}건 삭제', 'success')
    return redirect(url_for('vehicle_master', center=center_code))


@app.route('/masters/vehicle/upload', methods=['POST'])
def vehicle_master_upload():
    center_code = request.form.get('center_code', '').strip()
    if not center_code:
        flash('출발 센터를 선택해주세요.', 'danger')
        return redirect(url_for('vehicle_master'))
    file = request.files.get('file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('vehicle_master', center=center_code))
    try:
        df = pd.read_excel(file) if file.filename.endswith(('.xlsx', '.xls')) else pd.read_csv(file)
        df.columns = [str(c).strip() for c in df.columns]

        dest_col = df.columns[0]
        for col in df.columns:
            if any(kw in col for kw in ['도착', '출발', '지역', '도시']):
                dest_col = col
                break

        known_vehicles = ['11톤', '5톤장축', '5톤', '3.5톤', '2.5톤', '1.4톤', '1톤', '퀵']
        vehicle_cols = []
        for col in df.columns:
            clean = col.replace('.1', '').replace('.2', '').strip()
            if clean in known_vehicles and clean not in [v for _, v in vehicle_cols]:
                vehicle_cols.append((col, clean))

        if not vehicle_cols:
            flash('차량 종류 컬럼을 찾지 못했습니다. (11톤, 5톤, 1톤 등)', 'danger')
            return redirect(url_for('vehicle_master', center=center_code))

        overwrite = request.form.get('overwrite') == 'yes'
        if overwrite:
            VehicleRate.query.filter_by(center_code=center_code).delete()

        added = updated = 0
        for _, row in df.iterrows():
            dest_val = row.get(dest_col)
            if pd.isna(dest_val) or not str(dest_val).strip():
                continue
            dest = str(dest_val).strip()
            for raw_col, vtype in vehicle_cols:
                price_val = row.get(raw_col)
                if pd.isna(price_val):
                    continue
                try:
                    price = int(str(price_val).replace(',', '').replace(' ', ''))
                except ValueError:
                    continue
                existing = VehicleRate.query.filter_by(
                    center_code=center_code, destination=dest, vehicle_type=vtype
                ).first()
                if existing:
                    existing.unit_price = price
                    updated += 1
                else:
                    db.session.add(VehicleRate(
                        center_code=center_code, destination=dest,
                        vehicle_type=vtype, unit_price=price
                    ))
                    added += 1

        db.session.commit()
        center_name = OurCenter.query.filter_by(center_code=center_code).first()
        center_name = center_name.center_name if center_name else center_code
        flash(f'[{center_name}] 차량 단가 업로드 — {added}건 추가, {updated}건 수정', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'업로드 오류: {e}', 'danger')
    return redirect(url_for('vehicle_master', center=center_code))


@app.route('/masters/vehicle/add', methods=['POST'])
def vehicle_master_add():
    center_code  = request.form.get('center_code', '').strip()
    destination  = request.form.get('destination', '').strip()
    vehicle_type = request.form.get('vehicle_type', '').strip()
    price_str    = request.form.get('unit_price', '').replace(',', '').strip()
    if not all([center_code, destination, vehicle_type, price_str]):
        flash('모든 항목을 입력해주세요.', 'danger')
        return redirect(url_for('vehicle_master', center=center_code))
    try:
        price = int(price_str)
        existing = VehicleRate.query.filter_by(
            center_code=center_code, destination=destination, vehicle_type=vehicle_type
        ).first()
        if existing:
            existing.unit_price = price
            flash(f'단가 수정 완료', 'success')
        else:
            db.session.add(VehicleRate(
                center_code=center_code, destination=destination,
                vehicle_type=vehicle_type, unit_price=price
            ))
            flash(f'단가 추가 완료', 'success')
        db.session.commit()
    except Exception as e:
        flash(f'오류: {e}', 'danger')
    return redirect(url_for('vehicle_master', center=center_code))


@app.route('/masters/vehicle/<int:vid>/delete', methods=['POST'])
def vehicle_master_delete(vid):
    v = VehicleRate.query.get_or_404(vid)
    center_code = v.center_code
    db.session.delete(v)
    db.session.commit()
    flash('삭제되었습니다.', 'warning')
    return redirect(url_for('vehicle_master', center=center_code))


@app.route('/masters/vehicle/clear', methods=['POST'])
def vehicle_master_clear():
    center_code = request.form.get('center_code', '').strip()
    if center_code:
        cnt = VehicleRate.query.filter_by(center_code=center_code).delete()
        db.session.commit()
        flash(f'[{center_code}] 차량 단가 {cnt}건 초기화', 'warning')
    else:
        VehicleRate.query.delete()
        db.session.commit()
        flash('전체 차량 단가가 초기화되었습니다.', 'warning')
    return redirect(url_for('vehicle_master', center=center_code))


@app.route('/masters/vehicle/capacity/update', methods=['POST'])
def vehicle_capacity_update():
    caps = VehicleCapacity.query.all()
    for cap in caps:
        val = request.form.get(f'cap_{cap.id}')
        if val:
            try:
                cap.max_plt = float(val)
            except ValueError:
                pass
    db.session.commit()
    flash('차량 적재량이 업데이트되었습니다.', 'success')
    center = request.form.get('center_code', '')
    return redirect(url_for('vehicle_master', center=center))


@app.route('/masters/vehicle/template')
def vehicle_master_template():
    data = {
        '도착지':  ['강원도 강릉시', '강원도 춘천시', '경기도 가평군', '경기도 고양시'],
        '11톤':    [334000, 291000, 269000, 258000],
        '5톤장축': [289000, 251000, 235000, 220000],
        '5톤':     [253000, 219000, 205000, 196000],
        '3.5톤':   [220000, 194000, 183000, 174000],
        '2.5톤':   [201000, 181000, 172000, 165000],
        '1.4톤':   [162000, 145000, 139000, 132000],
        '1톤':     [154000, 136000, 130000, 125000],
        '퀵':      [149000, 129000, 124000, 118000],
    }
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='차량단가마스터')
    output.seek(0)
    return send_file(output, download_name='차량단가마스터_템플릿.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── 상품 마스터 ──────────────────────────────────────────────────────────────

@app.route('/customers/<int:cid>/products')
def product_master(cid):
    customer = Customer.query.get_or_404(cid)
    products = ProductMaster.query.filter_by(customer_id=cid).order_by(ProductMaster.product_code).all()
    return render_template('masters/product.html', customer=customer, products=products)


@app.route('/customers/<int:cid>/products/upload', methods=['POST'])
def product_upload(cid):
    Customer.query.get_or_404(cid)
    file = request.files.get('file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('product_master', cid=cid))
    try:
        df = pd.read_excel(file) if file.filename.endswith(('.xlsx', '.xls')) else pd.read_csv(file)
        df.columns = [str(c).strip() for c in df.columns]

        # 컬럼 매핑 (실제 Excel 헤더명 → 내부 컬럼명)
        col_map = {
            '고객사상품코드': 'product_code', '상품코드': 'product_code',
            '상품명': 'product_name',
            'BOX입수': 'box_in_count',
            '박스가로': 'box_width', '박스세로': 'box_depth', '박스높이': 'box_height',
            '박스중량': 'box_weight_kg',
            'PLT입수': 'plt_per_box',
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

        if 'product_code' not in df.columns:
            flash('상품코드 컬럼을 찾지 못했습니다. (고객사상품코드 또는 상품코드)', 'danger')
            return redirect(url_for('product_master', cid=cid))

        overwrite = request.form.get('overwrite') == 'yes'
        if overwrite:
            ProductMaster.query.filter_by(customer_id=cid).delete()

        added, updated = 0, 0
        for _, row in df.iterrows():
            code_val = row.get('product_code')
            if pd.isna(code_val):
                continue
            code = str(code_val).strip()

            def get_float(col):
                v = row.get(col)
                return float(v) if col in df.columns and not pd.isna(v) else None

            def get_int(col):
                v = row.get(col)
                return int(float(v)) if col in df.columns and not pd.isna(v) else None

            existing = ProductMaster.query.filter_by(customer_id=cid, product_code=code).first()
            kwargs = dict(
                product_name=str(row.get('product_name', code)).strip(),
                box_in_count=get_int('box_in_count'),
                box_width=get_float('box_width'),
                box_depth=get_float('box_depth'),
                box_height=get_float('box_height'),
                box_weight_kg=get_float('box_weight_kg'),
                plt_per_box=get_int('plt_per_box'),
            )
            if existing and not overwrite:
                for k, v in kwargs.items():
                    if v is not None:
                        setattr(existing, k, v)
                updated += 1
            else:
                db.session.add(ProductMaster(customer_id=cid, product_code=code, **kwargs))
                added += 1
        db.session.commit()
        flash(f'상품마스터: {added}건 추가, {updated}건 업데이트', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'업로드 오류: {e}', 'danger')
    return redirect(url_for('product_master', cid=cid))


@app.route('/customers/<int:cid>/products/template')
def product_template(cid):
    df = pd.DataFrame({
        '고객사상품코드': ['P001', 'P002'],
        '상품명': ['상품A', '상품B'],
        'BOX입수': [6, 4],
        '박스가로': [27.0, 30.0],
        '박스세로': [20.3, 22.0],
        '박스높이': [8.9, 10.0],
        '박스중량': [2.9, 4.5],
        'PLT입수': [300, 200],
    })
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='상품마스터')
    output.seek(0)
    return send_file(output, download_name='상품마스터_템플릿.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── 점포 마스터 ──────────────────────────────────────────────────────────────

@app.route('/customers/<int:cid>/stores')
def store_master(cid):
    customer = Customer.query.get_or_404(cid)
    stores = StoreMaster.query.filter_by(customer_id=cid).order_by(StoreMaster.sido, StoreMaster.store_code).all()
    unmapped = sum(1 for s in stores if not s.destination)
    return render_template('masters/store.html', customer=customer, stores=stores, unmapped=unmapped)


@app.route('/customers/<int:cid>/stores/upload', methods=['POST'])
def store_upload(cid):
    Customer.query.get_or_404(cid)
    file = request.files.get('file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('store_master', cid=cid))
    try:
        df = pd.read_excel(file) if file.filename.endswith(('.xlsx', '.xls')) else pd.read_csv(file)
        df.columns = [str(c).strip() for c in df.columns]

        col_map = {
            '배송처코드': 'store_code', '점포코드': 'store_code',
            '배송처명': 'store_name', '점포명': 'store_name', '센터': 'store_name',
            '주소': 'address',
            '시도': 'sido', '시군구': 'sigungu',
            '운영센터': 'center_name',
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

        # 배송처코드가 없으면 점포명으로 대체
        if 'store_code' not in df.columns and 'store_name' in df.columns:
            df['store_code'] = df['store_name']

        if 'store_code' not in df.columns:
            flash('배송처코드 또는 점포코드 컬럼이 필요합니다.', 'danger')
            return redirect(url_for('store_master', cid=cid))

        overwrite = request.form.get('overwrite') == 'yes'
        if overwrite:
            StoreMaster.query.filter_by(customer_id=cid).delete()

        added, updated = 0, 0
        for _, row in df.iterrows():
            code_val = row.get('store_code')
            if pd.isna(code_val):
                continue
            code = str(code_val).strip()

            address = str(row.get('address', '')).strip() if 'address' in df.columns and not pd.isna(row.get('address', float('nan'))) else None
            sido_val = str(row.get('sido', '')).strip() if 'sido' in df.columns and not pd.isna(row.get('sido', float('nan'))) else None
            sigungu_val = str(row.get('sigungu', '')).strip() if 'sigungu' in df.columns and not pd.isna(row.get('sigungu', float('nan'))) else None

            # 주소에서 시도+시군구 추출
            if address and (not sido_val or not sigungu_val):
                parsed_sido, parsed_sgg = extract_sido_sigungu(address)
                sido_val = sido_val or (normalize_sido(parsed_sido) if parsed_sido else None)
                sigungu_val = sigungu_val or parsed_sgg

            # 도착지 자동 매핑
            destination = None
            if sido_val and sigungu_val:
                dest_key = make_destination_key(normalize_sido(sido_val), sigungu_val)
                rate = VehicleRate.query.filter_by(destination=dest_key).first()
                if rate:
                    destination = dest_key
                else:
                    # 부분 매핑
                    rate = VehicleRate.query.filter(
                        VehicleRate.destination.like(f'%{sigungu_val}%')
                    ).first()
                    if rate:
                        destination = rate.destination

            existing = StoreMaster.query.filter_by(customer_id=cid, store_code=code).first()
            kwargs = dict(
                store_name=str(row.get('store_name', code)).strip() if 'store_name' in df.columns and not pd.isna(row.get('store_name', float('nan'))) else code,
                address=address,
                sido=normalize_sido(sido_val) if sido_val else None,
                sigungu=sigungu_val,
                destination=destination,
                center_name=str(row.get('center_name', '')).strip() if 'center_name' in df.columns and not pd.isna(row.get('center_name', float('nan'))) else None,
            )
            if existing and not overwrite:
                for k, v in kwargs.items():
                    if v is not None:
                        setattr(existing, k, v)
                updated += 1
            else:
                db.session.add(StoreMaster(customer_id=cid, store_code=code, **kwargs))
                added += 1
        db.session.commit()
        total = StoreMaster.query.filter_by(customer_id=cid).count()
        mapped = StoreMaster.query.filter(
            StoreMaster.customer_id == cid,
            StoreMaster.destination.isnot(None)
        ).count()
        flash(f'점포마스터: {added}건 추가, {updated}건 업데이트 | 도착지 자동매핑: {mapped}/{total}개', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'업로드 오류: {e}', 'danger')
    return redirect(url_for('store_master', cid=cid))


@app.route('/customers/<int:cid>/stores/template')
def store_template(cid):
    df = pd.DataFrame({
        '배송처코드': ['4630267', '1234567'],
        '배송처명': ['㈜이마트 보정몰센터', '㈜롯데마트 강릉점'],
        '주소': ['경기 용인시 기흥구 용구대로 2467', '강원도 강릉시 하슬라로 123'],
        '운영센터': ['이천센터', '이천센터'],
    })
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='점포마스터')
    output.seek(0)
    return send_file(output, download_name='점포마스터_템플릿.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── 출고내역 업로드 ──────────────────────────────────────────────────────────

@app.route('/customers/<int:cid>/calculate', methods=['GET'])
def calculate_page(cid):
    from calculator import get_direct_plt_threshold
    customer = Customer.query.get_or_404(cid)
    batches = db.session.query(
        ShippingHistory.batch_id,
        db.func.count(ShippingHistory.id).label('cnt'),
        db.func.min(ShippingHistory.uploaded_at).label('uploaded_at')
    ).filter_by(customer_id=cid).group_by(ShippingHistory.batch_id).order_by(
        db.func.min(ShippingHistory.uploaded_at).desc()
    ).all()
    threshold = get_direct_plt_threshold(db.session)
    centers = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).all()
    synergy_cnt = SynergyRoute.query.count()
    return render_template('calculation/index.html',
                           customer=customer, batches=batches,
                           threshold=threshold, centers=centers,
                           synergy_cnt=synergy_cnt)


@app.route('/customers/<int:cid>/history/<batch_id>/delete', methods=['POST'])
def history_delete(cid, batch_id):
    ShippingHistory.query.filter_by(customer_id=cid, batch_id=batch_id).delete()
    db.session.commit()
    flash('출고내역 배치가 삭제되었습니다.', 'warning')
    return redirect(url_for('calculate_page', cid=cid))


@app.route('/customers/<int:cid>/history/upload', methods=['POST'])
def history_upload(cid):
    Customer.query.get_or_404(cid)
    file = request.files.get('file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('calculate_page', cid=cid))
    try:
        df = pd.read_excel(file) if file.filename.endswith(('.xlsx', '.xls')) else pd.read_csv(file)
        df.columns = [str(c).strip() for c in df.columns]

        col_map = {
            '납품일자': 'shipping_date', '출고일': 'shipping_date', '일자': 'shipping_date',
            '주문번호': 'order_no', '오더유형': 'order_type', '채널': 'channel',
            '배송처코드': 'store_code', '점포코드': 'store_code', '거래처코드': 'store_code',
            '배송처명': 'store_name', '점포명': 'store_name', '거래처명': 'store_name',
            '주소': 'address', '배송주소': 'address',
            '상품코드': 'product_code', '품목코드': 'product_code',
            '상품명': 'product_name', '품목명': 'product_name',
            '출고수량(BOX)': 'box_qty', '출고수량': 'box_qty', '수량(BOX)': 'box_qty', '수량': 'box_qty', '박스수': 'box_qty',
            '출고수량(PLT)': 'plt_qty_decimal',  # PLT 컬럼명 고정: 반드시 이 이름만 인식
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

        # 필수 컬럼 확인 및 사용자에게 피드백
        has_store = 'store_code' in df.columns or 'store_name' in df.columns
        has_box = 'box_qty' in df.columns
        has_plt = 'plt_qty_decimal' in df.columns

        if not has_store:
            orig_cols = list(df.columns)
            flash(f'배송처 컬럼을 찾을 수 없습니다. 파일의 컬럼: {orig_cols[:10]}', 'danger')
            return redirect(url_for('calculate_page', cid=cid))
        if not has_box:
            flash('박스 수량 컬럼(출고수량(BOX))을 찾을 수 없습니다. 템플릿을 확인해주세요.', 'warning')
        if not has_plt:
            flash('출고수량(PLT) 컬럼이 없습니다. 상품마스터(PLT입수)로 자동 계산합니다.', 'info')

        # 헬퍼 함수 (루프 밖에 정의)
        def safe_str(row, col):
            v = row.get(col)
            if v is None or (isinstance(v, float) and math.isnan(v)):
                return None
            s = str(v).strip()
            return s if s and s.lower() != 'nan' else None

        def safe_float(row, col):
            v = row.get(col)
            if v is None or (isinstance(v, float) and math.isnan(v)):
                return None
            try:
                return float(v)
            except Exception:
                return None

        batch_id = str(uuid.uuid4())[:8]
        added = 0
        for _, row in df.iterrows():
            sc = safe_str(row, 'store_code')
            sn = safe_str(row, 'store_name')
            if not sc and not sn:
                continue

            ship_date = None
            if 'shipping_date' in df.columns:
                sd = row.get('shipping_date')
                if sd is not None and not (isinstance(sd, float) and math.isnan(sd)):
                    try:
                        ship_date = pd.to_datetime(sd).date()
                    except Exception:
                        pass

            plt_dec = safe_float(row, 'plt_qty_decimal')

            db.session.add(ShippingHistory(
                customer_id=cid, batch_id=batch_id,
                shipping_date=ship_date,
                order_no=safe_str(row, 'order_no'),
                order_type=safe_str(row, 'order_type'),
                channel=safe_str(row, 'channel'),
                store_code=sc, store_name=sn,
                address=safe_str(row, 'address'),
                product_code=safe_str(row, 'product_code'),
                product_name=safe_str(row, 'product_name'),
                box_qty=safe_float(row, 'box_qty'),
                plt_qty_decimal=plt_dec,
                plt_qty_int=None,
            ))
            added += 1
        db.session.commit()
        flash(f'✅ {added:,}건 출고내역 업로드 완료', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'업로드 오류: {e}', 'danger')
    return redirect(url_for('calculate_page', cid=cid))


@app.route('/customers/<int:cid>/history/template')
def history_template(cid):
    df = pd.DataFrame({
        '납품일자': ['2026-05-01', '2026-05-01'],
        '주문번호': ['F6486', 'F6487'],
        '오더유형': ['B2B', 'B2B'],
        '채널': ['할인점', '할인점'],
        '배송처코드': ['4630267', '4630267'],
        '배송처명': ['㈜이마트 보정몰센터', '㈜이마트 보정몰센터'],
        '주소': ['경기 용인시 기흥구 용구대로 2467', '경기 용인시 기흥구 용구대로 2467'],
        '상품코드': ['PG1011142056', 'PG1074021083'],
        '상품명': ['상품A', '상품B'],
        '출고수량(BOX)': [10, 5],
        '출고수량(PLT)': [0.5, 0.25],
    })
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='출고내역')
    output.seek(0)
    return send_file(output, download_name='출고내역_템플릿.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── 단가 산정 실행 ───────────────────────────────────────────────────────────

@app.route('/customers/<int:cid>/calculate/run', methods=['POST'])
def calculate_run(cid):
    customer = Customer.query.get_or_404(cid)
    batch_id          = request.form.get('batch_id')
    main_center_code  = request.form.get('main_center_code', '').strip()
    calc_name = request.form.get('calc_name', f'{customer.name} 단가산정 {datetime.now().strftime("%Y-%m-%d")}')

    if not main_center_code:
        flash('메인 센터를 선택해주세요.', 'danger')
        return redirect(url_for('calculate_page', cid=cid))

    query = ShippingHistory.query.filter_by(customer_id=cid)
    if batch_id:
        query = query.filter_by(batch_id=batch_id)
    history_rows = query.all()

    if not history_rows:
        flash('계산할 출고내역이 없습니다.', 'danger')
        return redirect(url_for('calculate_page', cid=cid))

    if not VehicleRate.query.filter_by(center_code=main_center_code).count():
        flash('선택한 센터의 차량 단가 마스터가 없습니다. 차량 단가를 먼저 등록해주세요.', 'danger')
        return redirect(url_for('calculate_page', cid=cid))

    results, errors, _ = calculate_from_history(history_rows, cid, calc_name, main_center_code, db.session)

    result_batch = str(uuid.uuid4())[:8]
    for r in results:
        db.session.add(CalculationResult(
            customer_id=cid,
            calc_name=calc_name,
            batch_id=result_batch,
            shipping_date=r['shipping_date'],
            store_code=r['store_code'],
            store_name=r['store_name'],
            address=r['address'],
            destination=r['destination'],
            delivery_mode=r['delivery_mode'],
            total_box_qty=r['total_box_qty'],
            total_plt_decimal=r['total_plt_decimal'],
            total_plt_count=r['total_plt_count'],
            vehicle_type=r['vehicle_type'],
            delivery_cost=r['delivery_cost'],
            transfer_cost=r.get('transfer_cost'),
            hub_cost=r.get('hub_cost'),
            cost_per_box=r['cost_per_box'],
            memo=r['memo'],
        ))
    db.session.commit()

    summary = summarize_results(results)

    if errors:
        flash(f'{len(errors)}건 처리 이슈 (단가 미산출)', 'warning')

    return render_template('calculation/result.html',
                           customer=customer,
                           results=results,
                           summary=summary,
                           errors=errors,
                           calc_name=calc_name,
                           result_batch=result_batch)


# ─── 결과 조회 & 내보내기 ────────────────────────────────────────────────────────

@app.route('/customers/<int:cid>/results')
def result_list(cid):
    customer = Customer.query.get_or_404(cid)
    results = CalculationResult.query.filter_by(customer_id=cid).order_by(
        CalculationResult.calc_date.desc()
    ).all()
    batches = {}
    for r in results:
        bid = r.batch_id
        if bid not in batches:
            batches[bid] = {'name': r.calc_name, 'date': r.calc_date, 'rows': []}
        batches[bid]['rows'].append(r)
    return render_template('customers/results.html', customer=customer, batches=batches)


@app.route('/customers/<int:cid>/results/<batch_id>/export')
def result_export(cid, batch_id):
    customer = Customer.query.get_or_404(cid)
    results = CalculationResult.query.filter_by(customer_id=cid, batch_id=batch_id).all()
    if not results:
        flash('결과가 없습니다.', 'danger')
        return redirect(url_for('result_list', cid=cid))

    rows = [{
        '납품일자': r.shipping_date,
        '배송처코드': r.store_code,
        '배송처명': r.store_name,
        '주소': r.address,
        '도착지': r.destination,
        '배송모드': r.delivery_mode,
        '박스수': r.total_box_qty,
        'PLT수(소수)': r.total_plt_decimal,
        'PLT수(올림)': r.total_plt_count,
        '차량종류': r.vehicle_type,
        '배송비(원)': r.delivery_cost,
        '박스당배송비(원)': r.cost_per_box,
        '비고': r.memo,
    } for r in results]

    df = pd.DataFrame(rows)

    # 요약 행
    valid = [r for r in results if r.delivery_cost]
    total_cost = sum(r.delivery_cost for r in valid)
    total_boxes = sum(r.total_box_qty or 0 for r in valid)
    avg_cpb = round(total_cost / total_boxes, 1) if total_boxes > 0 else None
    df = pd.concat([df, pd.DataFrame([{
        '납품일자': '합계/평균', '배송처코드': '', '배송처명': '',
        '주소': '', '도착지': '', '배송모드': '',
        '박스수': total_boxes,
        'PLT수(소수)': sum(r.total_plt_decimal or 0 for r in valid),
        'PLT수(올림)': sum(r.total_plt_count or 0 for r in valid),
        '차량종류': '', '배송비(원)': total_cost, '박스당배송비(원)': avg_cpb, '비고': '',
    }])], ignore_index=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='단가산정결과')
    output.seek(0)
    filename = f'{customer.name}_배송단가_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/customers/<int:cid>/results/<batch_id>/delete', methods=['POST'])
def result_delete(cid, batch_id):
    CalculationResult.query.filter_by(customer_id=cid, batch_id=batch_id).delete()
    db.session.commit()
    flash('결과가 삭제되었습니다.', 'warning')
    return redirect(url_for('result_list', cid=cid))


# ─── 주소 좌표 변환 API ────────────────────────────────────────────────────────

@app.route('/api/geocode', methods=['POST'])
def api_geocode():
    address = (request.json or {}).get('address', '').strip()
    if not address:
        return jsonify({'error': '주소를 입력해주세요'}), 400

    api_key_cfg = SystemConfig.query.filter_by(key='kakao_api_key').first()
    api_key = api_key_cfg.value.strip() if api_key_cfg else ''
    if not api_key:
        return jsonify({'error': '카카오 API 키가 설정되지 않았습니다'}), 500

    try:
        resp = http_req.get(
            'https://dapi.kakao.com/v2/local/search/address.json',
            headers={'Authorization': f'KakaoAK {api_key}'},
            params={'query': address, 'size': 1},
            timeout=5,
        )
        docs = resp.json().get('documents', [])
        if docs:
            return jsonify({'lat': float(docs[0]['y']), 'lon': float(docs[0]['x']),
                            'address_name': docs[0].get('address_name', address)})
        return jsonify({'error': '주소를 찾을 수 없습니다'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── 시스템 설정 ──────────────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    from calculator import get_direct_plt_threshold
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'threshold':
            threshold = request.form.get('direct_plt_threshold', '').strip()
            try:
                val = float(threshold)
                cfg = SystemConfig.query.filter_by(key='direct_plt_threshold').first()
                if cfg:
                    cfg.value = str(val)
                else:
                    db.session.add(SystemConfig(key='direct_plt_threshold', value=str(val),
                                                description='직송 전환 기준 PLT 수'))
                db.session.commit()
                flash(f'직송 기준이 PLT {val}개 이상으로 변경되었습니다.', 'success')
            except ValueError:
                flash('숫자를 입력해주세요.', 'danger')

        elif action == 'api_keys':
            kakao_key = request.form.get('kakao_api_key', '').strip()
            stops = request.form.get('stops_per_vehicle', '').strip()
            for key, val, desc in [
                ('kakao_api_key', kakao_key, '카카오 주소→좌표 변환 REST API 키'),
                ('stops_per_vehicle', stops, '공동배송 차량당 평균 도착지(점포) 수'),
            ]:
                if val:
                    cfg = SystemConfig.query.filter_by(key=key).first()
                    if cfg:
                        cfg.value = val
                    else:
                        db.session.add(SystemConfig(key=key, value=val, description=desc))
            db.session.commit()
            flash('API 키 및 공동배송 설정이 저장되었습니다.', 'success')

        return redirect(url_for('settings'))

    current_threshold = get_direct_plt_threshold(db.session)
    kakao_key = (SystemConfig.query.filter_by(key='kakao_api_key').first() or type('', (), {'value': ''})()).value
    stops_per_vehicle = int((SystemConfig.query.filter_by(key='stops_per_vehicle').first() or type('', (), {'value': '8'})()).value)
    return render_template('settings.html',
                           threshold=current_threshold,
                           kakao_api_key=kakao_key,
                           stops_per_vehicle=stops_per_vehicle)


# ─── 공동배송 단가 관리 ────────────────────────────────────────────────────────

@app.route('/masters/joint')
def joint_master():
    from calculator import get_direct_plt_threshold
    rates = JointDeliveryRate.query.order_by(JointDeliveryRate.destination).all()
    destinations = db.session.query(VehicleRate.destination).distinct().order_by(VehicleRate.destination).all()
    destinations = ['기본'] + [d[0] for d in destinations]
    threshold = get_direct_plt_threshold(db.session)
    return render_template('masters/joint.html', rates=rates, destinations=destinations, threshold=threshold)


@app.route('/masters/joint/add', methods=['POST'])
def joint_rate_add():
    destination = request.form.get('destination', '').strip()
    price_str = request.form.get('price_per_box', '').replace(',', '').strip()
    memo = request.form.get('memo', '').strip()
    if not destination or not price_str:
        flash('도착지와 단가를 입력해주세요.', 'danger')
        return redirect(url_for('joint_master'))
    try:
        price = float(price_str)
        existing = JointDeliveryRate.query.filter_by(destination=destination).first()
        if existing:
            existing.price_per_box = price
            existing.memo = memo
            flash(f'[{destination}] 공동배송 단가가 업데이트되었습니다.', 'success')
        else:
            db.session.add(JointDeliveryRate(destination=destination, price_per_box=price, memo=memo))
            flash(f'[{destination}] 공동배송 단가가 추가되었습니다.', 'success')
        db.session.commit()
    except Exception as e:
        flash(f'오류: {e}', 'danger')
    return redirect(url_for('joint_master'))


@app.route('/masters/joint/<int:rid>/delete', methods=['POST'])
def joint_rate_delete(rid):
    r = JointDeliveryRate.query.get_or_404(rid)
    name = r.destination
    db.session.delete(r)
    db.session.commit()
    flash(f'[{name}] 단가가 삭제되었습니다.', 'warning')
    return redirect(url_for('joint_master'))


@app.route('/masters/joint/upload', methods=['POST'])
def joint_rate_upload():
    file = request.files.get('file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('joint_master'))
    try:
        df = pd.read_excel(file) if file.filename.endswith(('.xlsx', '.xls')) else pd.read_csv(file)
        df.columns = [str(c).strip() for c in df.columns]
        col_map = {'도착지': 'destination', '박스당단가': 'price_per_box', '단가': 'price_per_box', '메모': 'memo'}
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
        if 'destination' not in df.columns or 'price_per_box' not in df.columns:
            flash('필수 컬럼: 도착지, 박스당단가', 'danger')
            return redirect(url_for('joint_master'))
        added, updated = 0, 0
        for _, row in df.iterrows():
            if pd.isna(row.get('destination')):
                continue
            dest = str(row['destination']).strip()
            price = float(row['price_per_box'])
            memo = str(row.get('memo', '')).strip() if 'memo' in df.columns and not pd.isna(row.get('memo')) else None
            existing = JointDeliveryRate.query.filter_by(destination=dest).first()
            if existing:
                existing.price_per_box = price
                existing.memo = memo
                updated += 1
            else:
                db.session.add(JointDeliveryRate(destination=dest, price_per_box=price, memo=memo))
                added += 1
        db.session.commit()
        flash(f'공동배송 단가: {added}건 추가, {updated}건 업데이트', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'업로드 오류: {e}', 'danger')
    return redirect(url_for('joint_master'))


# ─── 자사 센터 관리 ────────────────────────────────────────────────────────────

INITIAL_CENTERS = [
    # (코드, 명칭, 주소, 위도, 경도, 직송허브여부, 정렬순서)
    ('3300',    '인천센터',       '인천광역시 서구 북항로245번길 13-22',           37.50399, 126.62355, True,  0),
    ('3100',    '인천대포',       '인천광역시 서구 북항로 178번길 68-47',          37.51327, 126.63373, True,  1),
    ('3500',    '화성센터본점',   '경기도 화성시 양감면 초록로 103',               37.07786, 126.95455, True,  2),
    ('2000',    '이천센터',       '경기도 이천시 대월면 대월로 331',               37.18854, 127.49480, True,  3),
    ('2100',    '이천R2',         '경기도 이천시 대월면 대월로 331',               37.18854, 127.49480, True,  4),
    ('2200',    '이천대포',       '경기도 이천시 대월면 대월로932번길 63',         37.23642, 127.50253, True,  5),
    ('1100D',   '대월센터',       '경기도 이천시 대월면 대월로 627-61',            37.21011, 127.49390, True,  6),
    ('200',     '남이천1센터',    '경기도 이천시 모가면 공원로 112',               37.18209, 127.45063, True,  7),
    ('100',     '남이천2센터',    '경기도 이천시 모가면 공원로 134',               37.18194, 127.45288, True,  8),
    ('1100',    '동이천센터',     '경기도 이천시 부발읍 황무로 2037-57',           37.24314, 127.51581, True,  9),
    ('2700',    '북이천센터',     '경기도 이천시 부발읍 중부대로 2051번길 101',    37.29595, 127.52966, True, 10),
    ('5400',    '백암센터',       '경기도 용인시 처인구 백암면 덕평로 120',        37.18022, 127.37364, True, 11),
    ('7000',    '광주오포센터',   '경기도 광주시 오포로 614',                      37.35163, 127.21486, True, 12),
    ('9900',    '진천센터',       '충청북도 진천군 초평면 용정길 29-7',            36.84157, 127.51628, True, 13),
    ('0090000', '남청주센터',     '청주시 서원구 남이면 저산척북로 203-30',        36.56295, 127.38343, True, 14),
    ('5000',    '대구센터',       '경상북도 경산시 압량읍 가야로 20',              35.85336, 128.76886, True, 15),
    ('6000',    '김해센터',       '경상남도 김해시 상동면 상동로 680-70',          35.31140, 128.93961, True, 16),
]


def _seed_centers():
    if OurCenter.query.count() == 0:
        for row in INITIAL_CENTERS:
            code, name, addr, lat, lon, is_hub, sort = row
            db.session.add(OurCenter(
                center_code=code, center_name=name, address=addr,
                lat=lat, lon=lon, is_direct_hub=is_hub, sort_order=sort
            ))
        db.session.commit()


# ── 이고비용 기본 데이터 ──────────────────────────────────────────────────────
# 차량 1대 전체 운행 단가 (원). 실제 이고비 = unit_price × (실제PLT / 차량최대PLT)
# 거리 등급: 근(~70km) / 중(70~200km) / 원(200km+)
_TR_PRICES = {
    '근': {'11톤': 350000, '5톤장축': 250000, '5톤': 200000, '3.5톤': 120000, '2.5톤': 95000},
    '중': {'11톤': 550000, '5톤장축': 400000, '5톤': 320000, '3.5톤': 195000, '2.5톤': 160000},
    '원': {'11톤': 850000, '5톤장축': 630000, '5톤': 500000, '3.5톤': 310000, '2.5톤': 255000},
}
_TRANSFER_ROUTES = [
    # (출발, 도착, 거리등급)  —  실제 운영 노선 기준
    # 인천센터(3300) 출발
    ('3300', '3100',    '근'), ('3300', '3500',    '근'), ('3300', '2000',    '중'),
    ('3300', '7000',    '중'), ('3300', '5400',    '중'), ('3300', '9900',    '중'),
    ('3300', '0090000', '중'), ('3300', '5000',    '원'), ('3300', '6000',    '원'),
    # 이천센터(2000) 출발
    ('2000', '3300',    '중'), ('2000', '3500',    '중'), ('2000', '7000',    '근'),
    ('2000', '5400',    '근'), ('2000', '9900',    '중'), ('2000', '0090000', '중'),
    ('2000', '5000',    '원'), ('2000', '6000',    '원'),
    # 이천 내부 (대포·서브센터)
    ('2000', '2200',    '근'), ('2000', '1100D',   '근'), ('2000', '200',     '근'),
    ('2000', '100',     '근'), ('2000', '1100',    '근'), ('2000', '2700',    '근'),
    ('2000', '2100',    '근'),
    # 화성센터(3500) 출발
    ('3500', '3300',    '근'), ('3500', '2000',    '중'), ('3500', '0090000', '중'),
    # 광주오포(7000) 출발
    ('7000', '2000',    '근'), ('7000', '5400',    '근'),
    # 진천(9900), 남청주(0090000) 출발
    ('9900',    '0090000', '근'), ('9900', '5000', '중'), ('9900', '6000', '원'),
    ('0090000', '5000',    '중'), ('0090000', '6000', '원'),
    # 김해(6000) ↔ 대구(5000)
    ('6000', '5000', '중'), ('5000', '6000', '중'),
]

INITIAL_TRANSFER_RATES = []
for _from, _to, _tier in _TRANSFER_ROUTES:
    for _vt, _price in _TR_PRICES[_tier].items():
        INITIAL_TRANSFER_RATES.append((_from, _to, _vt, _price))


# ── 거점 변동용차 기본 데이터 ─────────────────────────────────────────────────
# 차량 1대 전체 운행 단가. 실제 비용 = unit_price × (실적PLT / 차량최대PLT)
# (센터코드, 차량종류, 기본단가) — 배송지구 구분 없이 거점센터 단위로 관리
INITIAL_HUB_VEHICLE_RATES = [
    # (센터코드, 차량종류, 1회단가) -- 기본값, 실제 운용 후 수정 필요
    ('3300', '1톤',    90000),  ('3300', '2.5톤', 165000), ('3300', '5톤',    310000),
    ('3100', '1톤',    90000),  ('3100', '2.5톤', 165000), ('3100', '5톤',    310000),
    ('3500', '1톤',    95000),  ('3500', '2.5톤', 175000), ('3500', '5톤',    330000),
    ('2000', '1톤',    85000),  ('2000', '2.5톤', 155000), ('2000', '5톤',    290000),
    ('2100', '1톤',    85000),  ('2100', '2.5톤', 155000), ('2100', '5톤',    290000),
    ('2200', '1톤',    85000),  ('2200', '2.5톤', 155000), ('2200', '5톤',    290000),
    ('1100D','1톤',    85000),  ('1100D','2.5톤', 155000), ('1100D','5톤',    290000),
    ('200',  '1톤',    80000),  ('200',  '2.5톤', 148000), ('200',  '5톤',    278000),
    ('100',  '1톤',    80000),  ('100',  '2.5톤', 148000), ('100',  '5톤',    278000),
    ('1100', '1톤',    85000),  ('1100', '2.5톤', 155000), ('1100', '5톤',    290000),
    ('2700', '1톤',    90000),  ('2700', '2.5톤', 165000), ('2700', '5톤',    310000),
    ('5400', '1톤',    90000),  ('5400', '2.5톤', 165000), ('5400', '5톤',    310000),
    ('7000', '1톤',    85000),  ('7000', '2.5톤', 155000), ('7000', '5톤',    290000),
    ('9900', '1톤',    80000),  ('9900', '2.5톤', 148000), ('9900', '5톤',    278000),
    ('0090000','1톤',  80000),  ('0090000','2.5톤',148000),('0090000','5톤',  278000),
    ('5000', '1톤',    85000),  ('5000', '2.5톤', 155000), ('5000', '5톤',    290000),
    ('6000', '1톤',    85000),  ('6000', '2.5톤', 155000), ('6000', '5톤',    290000),
]


def _seed_transfer_and_hub():
    if TransferRate.query.count() == 0:
        for from_c, to_c, vt, price in INITIAL_TRANSFER_RATES:
            db.session.add(TransferRate(
                from_center_code=from_c, to_center_code=to_c,
                vehicle_type=vt, unit_price=price
            ))
    if HubVehicleRate.query.count() == 0:
        for code, vtype, price in INITIAL_HUB_VEHICLE_RATES:
            db.session.add(HubVehicleRate(
                center_code=code, vehicle_type=vtype, unit_price=price
            ))
    db.session.commit()


with app.app_context():
    _seed_centers()
    db.create_all()
    _seed_transfer_and_hub()


@app.route('/centers')
def center_list():
    centers = OurCenter.query.order_by(OurCenter.sort_order, OurCenter.center_name).all()
    return render_template('centers/list.html', centers=centers)


@app.route('/centers/add', methods=['POST'])
def center_add():
    code = request.form.get('center_code', '').strip().upper()
    name = request.form.get('center_name', '').strip()
    addr = request.form.get('address', '').strip()
    lat_s = request.form.get('lat', '').strip()
    lon_s = request.form.get('lon', '').strip()
    is_hub = request.form.get('is_direct_hub') == 'on'
    memo = request.form.get('memo', '').strip()
    if not code or not name:
        flash('센터코드와 센터명은 필수입니다.', 'danger')
        return redirect(url_for('center_list'))
    try:
        lat = float(lat_s) if lat_s else None
        lon = float(lon_s) if lon_s else None
        existing = OurCenter.query.filter_by(center_code=code).first()
        if existing:
            existing.center_name = name
            existing.address = addr
            existing.lat = lat
            existing.lon = lon
            existing.is_direct_hub = is_hub
            existing.memo = memo
            flash(f'[{code}] {name} 업데이트되었습니다.', 'success')
        else:
            max_order = db.session.query(db.func.max(OurCenter.sort_order)).scalar() or 0
            db.session.add(OurCenter(
                center_code=code, center_name=name, address=addr,
                lat=lat, lon=lon, is_direct_hub=is_hub, memo=memo,
                sort_order=max_order + 1
            ))
            flash(f'[{code}] {name} 추가되었습니다.', 'success')
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'오류: {e}', 'danger')
    return redirect(url_for('center_list'))


@app.route('/centers/<int:ctr_id>/edit', methods=['POST'])
def center_edit(ctr_id):
    center = OurCenter.query.get_or_404(ctr_id)
    name   = request.form.get('center_name', '').strip()
    addr   = request.form.get('address', '').strip()
    lat_s  = request.form.get('lat', '').strip()
    lon_s  = request.form.get('lon', '').strip()
    memo   = request.form.get('memo', '').strip()
    is_hub  = request.form.get('is_direct_hub') == 'on'
    is_main = request.form.get('is_main_center') == 'on'
    sort_s  = request.form.get('sort_order', '').strip()
    if not name:
        flash('센터명은 필수입니다.', 'danger')
        return redirect(url_for('center_list'))
    try:
        center.center_name   = name
        center.address       = addr
        center.lat           = float(lat_s) if lat_s else None
        center.lon           = float(lon_s) if lon_s else None
        center.memo          = memo
        center.is_direct_hub  = is_hub
        center.is_main_center = is_main
        if sort_s:
            center.sort_order = int(sort_s)
        db.session.commit()
        flash(f'[{center.center_code}] {name} 수정되었습니다.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'오류: {e}', 'danger')
    return redirect(url_for('center_list'))


@app.route('/centers/<int:ctr_id>/delete', methods=['POST'])
def center_delete(ctr_id):
    center = OurCenter.query.get_or_404(ctr_id)
    name = center.center_name
    db.session.delete(center)
    db.session.commit()
    flash(f'[{name}] 센터가 삭제되었습니다.', 'warning')
    return redirect(url_for('center_list'))


@app.route('/centers/<int:ctr_id>/toggle-hub', methods=['POST'])
def center_toggle_hub(ctr_id):
    center = OurCenter.query.get_or_404(ctr_id)
    center.is_direct_hub = not center.is_direct_hub
    db.session.commit()
    status = '직송허브' if center.is_direct_hub else '거점전용'
    flash(f'[{center.center_name}] → {status}로 변경되었습니다.', 'success')
    return redirect(url_for('center_list'))


@app.route('/centers/<int:ctr_id>/toggle-main', methods=['POST'])
def center_toggle_main(ctr_id):
    center = OurCenter.query.get_or_404(ctr_id)
    center.is_main_center = not center.is_main_center
    db.session.commit()
    status = '메인센터' if center.is_main_center else '일반센터'
    flash(f'[{center.center_name}] → {status}로 변경되었습니다.', 'success')
    return redirect(url_for('center_list'))


# ─── 이고비용 마스터 (매트릭스 뷰) ─────────────────────────────────────────────

@app.route('/masters/transfer')
def transfer_master():
    rates = TransferRate.query.all()
    # matrix[from_code][to_code][vt] = unit_price
    matrix = {}
    for r in rates:
        matrix.setdefault(r.from_center_code, {}).setdefault(r.to_center_code, {})[r.vehicle_type] = r.unit_price
    main_centers = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).all()
    all_centers  = OurCenter.query.order_by(OurCenter.sort_order).all()
    vehicle_types = [c.vehicle_type for c in VehicleCapacity.query.order_by(VehicleCapacity.sort_order).all()]
    return render_template('masters/transfer.html',
                           main_centers=main_centers, centers=all_centers,
                           vehicle_types=vehicle_types, matrix=matrix)


@app.route('/masters/transfer/save-matrix', methods=['POST'])
def transfer_matrix_save():
    from_code = request.form.get('from_code', '').strip()
    tos    = request.form.getlist('to')
    vts    = request.form.getlist('vt')
    prices = request.form.getlist('price')
    if not from_code:
        flash('오류: 출발 센터 미지정', 'danger')
        return redirect(url_for('transfer_master'))
    saved = deleted = 0
    for to, vt, p_str in zip(tos, vts, prices):
        p_str = p_str.replace(',', '').strip()
        existing = TransferRate.query.filter_by(
            from_center_code=from_code, to_center_code=to, vehicle_type=vt
        ).first()
        if p_str and int(p_str) > 0:
            price = int(p_str)
            if existing:
                existing.unit_price = price
            else:
                db.session.add(TransferRate(
                    from_center_code=from_code, to_center_code=to,
                    vehicle_type=vt, unit_price=price
                ))
            saved += 1
        elif existing:
            db.session.delete(existing)
            deleted += 1
    db.session.commit()
    flash(f'이고비용 저장 완료 — {saved}건 저장, {deleted}건 삭제', 'success')
    return redirect(url_for('transfer_master'))


# ─── 거점 변동용차 비용 마스터 ────────────────────────────────────────────────

@app.route('/masters/hub-vehicle')
def hub_vehicle_master():
    centers       = OurCenter.query.order_by(OurCenter.sort_order).all()
    capacities    = VehicleCapacity.query.order_by(VehicleCapacity.sort_order).all()
    vehicle_types = [c.vehicle_type for c in capacities]

    rates = HubVehicleRate.query.all()
    # matrix[center_code][vt] = unit_price
    matrix = {}
    for r in rates:
        matrix.setdefault(r.center_code, {})[r.vehicle_type] = r.unit_price

    return render_template('masters/hub_vehicle.html',
                           centers=centers, vehicle_types=vehicle_types,
                           capacities=capacities, matrix=matrix)


@app.route('/masters/hub-vehicle/save', methods=['POST'])
def hub_vehicle_save():
    center_code = request.form.get('center_code', '').strip()
    vts    = request.form.getlist('vt')
    prices = request.form.getlist('price')
    if not center_code:
        flash('오류: 센터 미지정', 'danger')
        return redirect(url_for('hub_vehicle_master'))
    saved = deleted = 0
    for vt, p_str in zip(vts, prices):
        p_str = p_str.replace(',', '').strip()
        existing = HubVehicleRate.query.filter_by(
            center_code=center_code, vehicle_type=vt
        ).first()
        if p_str and int(p_str) > 0:
            price = int(p_str)
            if existing:
                existing.unit_price = price
            else:
                db.session.add(HubVehicleRate(
                    center_code=center_code, vehicle_type=vt, unit_price=price
                ))
            saved += 1
        elif existing:
            db.session.delete(existing)
            deleted += 1
    db.session.commit()
    flash(f'변동용차 단가 저장 완료 — {saved}건 저장, {deleted}건 삭제', 'success')
    return redirect(url_for('hub_vehicle_master'))


# ─── 거점 변동용차 거리별 단가 ────────────────────────────────────────────────

@app.route('/masters/distance-rate')
def distance_rate_master():
    from sqlalchemy import func as sqlfunc
    from hub_rate_data import VEHICLE_TYPES, VEHICLE_PLT_CAP

    count = VehicleDistanceRate.query.count()
    # 각 차종별 min/max 단가 요약
    summary = {}
    for vt in VEHICLE_TYPES:
        rows = VehicleDistanceRate.query.filter_by(vehicle_type=vt).order_by(VehicleDistanceRate.km).all()
        if rows:
            summary[vt] = {
                'count': len(rows),
                'min_km': rows[0].km, 'max_km': rows[-1].km,
                'min_price': rows[0].unit_price, 'max_price': rows[-1].unit_price,
                'plt_cap': VEHICLE_PLT_CAP.get(vt, '-'),
            }

    # 샘플 데이터: 10km 단위로 전체 표 (100행 × 5열)
    sample_kms = list(range(1, 1001, 10))
    sample_rows = []
    rate_map = {}
    for vdr in VehicleDistanceRate.query.filter(VehicleDistanceRate.km.in_(sample_kms)).order_by(VehicleDistanceRate.km).all():
        rate_map.setdefault(vdr.km, {})[vdr.vehicle_type] = vdr.unit_price
    for km in sample_kms:
        if km in rate_map:
            sample_rows.append({'km': km, **rate_map[km]})

    return render_template('masters/distance_rate.html',
                           vehicle_types=VEHICLE_TYPES,
                           plt_cap=VEHICLE_PLT_CAP,
                           count=count,
                           summary=summary,
                           sample_rows=sample_rows)


@app.route('/masters/distance-rate/seed', methods=['POST'])
def distance_rate_seed():
    from hub_rate_data import DISTANCE_RATES, VEHICLE_TYPES
    try:
        VehicleDistanceRate.query.delete()
        rows = []
        for entry in DISTANCE_RATES:
            km = entry[0]
            for i, vtype in enumerate(VEHICLE_TYPES):
                rows.append(VehicleDistanceRate(vehicle_type=vtype, km=km, unit_price=entry[i + 1]))
        db.session.bulk_save_objects(rows)
        db.session.commit()
        flash(f'거리별 단가 {len(rows)}건 등록 완료 ({len(DISTANCE_RATES)}km x {len(VEHICLE_TYPES)}차종)', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'등록 실패: {e}', 'danger')
    return redirect(url_for('distance_rate_master'))


@app.route('/api/joint-breakdown/<int:customer_id>')
def api_joint_breakdown(customer_id):
    """이고비/변동용차비 거점별 상세 내역"""
    _spv_cfg = SystemConfig.query.filter_by(key='stops_per_vehicle').first()
    stops_per_vehicle = int(_spv_cfg.value) if _spv_cfg else 8
    _main_ctr = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).first()
    _main_code = _main_ctr.center_code if _main_ctr else None
    _lb = db.session.query(CalculationResult.batch_id).filter(
        CalculationResult.customer_id == customer_id
    ).order_by(CalculationResult.calc_date.desc()).first()
    _lb_id = _lb[0] if _lb else None
    detail = compute_joint_breakdown_detail(
        customer_id, _main_code, stops_per_vehicle, db.session, batch_id=_lb_id
    )
    detail['stops_per_vehicle'] = stops_per_vehicle
    return jsonify(detail)


@app.route('/api/threshold-preview/<int:customer_id>')
def api_threshold_preview(customer_id):
    """직송 기준 PLT 변경 시 직송/공동 비용 즉시 미리보기"""
    from calculator import compute_joint_breakdown_live
    threshold = request.args.get('threshold', type=float, default=3.0)

    main_ctr  = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).first()
    main_code = main_ctr.center_code if main_ctr else None

    _spv_cfg = SystemConfig.query.filter_by(key='stops_per_vehicle').first()
    stops_per_vehicle = int(_spv_cfg.value) if _spv_cfg else 8

    caps   = {r.vehicle_type: r.max_plt for r in VehicleCapacity.query.order_by(VehicleCapacity.sort_order).all()}
    vr_map = {}
    for vr in VehicleRate.query.filter_by(center_code=main_code).all():
        vr_map.setdefault(vr.destination, {})[vr.vehicle_type] = vr.unit_price
    vt_order = [r.vehicle_type for r in VehicleCapacity.query.order_by(VehicleCapacity.sort_order).all()]

    def vehicle_cost(plt_dec, destination):
        rates = vr_map.get(destination or '', {})
        if not rates:
            return 0
        best = None
        for vt in vt_order:
            if vt not in rates or vt not in caps:
                continue
            mp   = caps[vt]
            up   = rates[vt]
            cost = up if plt_dec <= mp else up * math.ceil(plt_dec / mp)
            if best is None or cost < best:
                best = cost
        return best or 0

    _lb2 = db.session.query(CalculationResult.batch_id).filter(
        CalculationResult.customer_id == customer_id
    ).order_by(CalculationResult.calc_date.desc()).first()
    _lb2_id = _lb2[0] if _lb2 else None
    _rec_f = CalculationResult.customer_id == customer_id
    if _lb2_id:
        _rec_f = _rec_f & (CalculationResult.batch_id == _lb2_id)
    records = CalculationResult.query.filter(_rec_f).all()

    direct_cost = direct_cnt = joint_cnt = 0
    direct_boxes = joint_boxes = 0
    direct_days = set()
    joint_days  = set()
    for r in records:
        plt = r.total_plt_decimal or 0
        if plt >= threshold:
            direct_cost  += vehicle_cost(plt, r.destination)
            direct_cnt   += 1
            direct_boxes += int(r.total_box_qty or 0)
            if r.shipping_date:
                direct_days.add(r.shipping_date)
        else:
            joint_cnt  += 1
            joint_boxes += int(r.total_box_qty or 0)
            if r.shipping_date:
                joint_days.add(r.shipping_date)

    # 공동배송 비용: threshold 기반으로 재분류 후 이고비 + 변동용차비 재계산
    joint_days_cnt = len(joint_days) or 1
    if joint_cnt > 0:
        transfer_cost, _ = compute_joint_breakdown_live(
            customer_id, main_code, stops_per_vehicle, db.session, threshold=threshold
        )
        hub_vehicle_cost = _hub_vehicle_daily_cost(customer_id, stops_per_vehicle, threshold=threshold) * joint_days_cnt
        joint_cost = transfer_cost + hub_vehicle_cost
    else:
        transfer_cost    = 0
        hub_vehicle_cost = 0
        joint_cost       = 0

    total_cost = direct_cost + joint_cost

    return jsonify({
        'threshold':       threshold,
        'direct_cost':     direct_cost,
        'direct_cnt':      direct_cnt,
        'direct_boxes':    direct_boxes,
        'joint_cost':      joint_cost,
        'transfer_cost':   transfer_cost,
        'hub_vehicle_cost': hub_vehicle_cost,
        'joint_cnt':       joint_cnt,
        'joint_boxes':     joint_boxes,
        'total_cost':      total_cost,
        'direct_days':     len(direct_days),
        'joint_days':      joint_days_cnt,
    })


@app.route('/api/direct-breakdown/<int:customer_id>')
def api_direct_breakdown(customer_id):
    """직송 물류비 도착지별 상세 내역 (threshold 파라미터로 실시간 재분류 가능)"""
    threshold = request.args.get('threshold', type=float)

    if threshold is None:
        # threshold 미지정 → DB의 delivery_mode 기준 (저장된 분류 그대로)
        rows = (
            db.session.query(
                CalculationResult.destination,
                CalculationResult.vehicle_type,
                db.func.count(CalculationResult.id).label('cnt'),
                db.func.sum(CalculationResult.total_plt_decimal).label('total_plt'),
                db.func.sum(CalculationResult.total_box_qty).label('total_box'),
                db.func.sum(CalculationResult.delivery_cost).label('total_cost'),
            )
            .filter(CalculationResult.customer_id == customer_id,
                    CalculationResult.delivery_mode == '직송')
            .group_by(CalculationResult.destination, CalculationResult.vehicle_type)
            .order_by(db.func.sum(CalculationResult.delivery_cost).desc())
            .all()
        )
        result = [{'destination': r.destination or '-', 'vehicle_type': r.vehicle_type or '-',
                   'cnt': int(r.cnt or 0), 'total_plt': round(float(r.total_plt or 0), 2),
                   'total_box': round(float(r.total_box or 0)),
                   'total_cost': int(r.total_cost or 0)} for r in rows]
    else:
        # threshold 지정 → PLT 기준으로 동적 재분류 후 vehicle_cost 계산
        main_ctr  = OurCenter.query.filter_by(is_main_center=True).order_by(OurCenter.sort_order).first()
        main_code = main_ctr.center_code if main_ctr else None
        caps    = {r.vehicle_type: r.max_plt for r in VehicleCapacity.query.order_by(VehicleCapacity.sort_order).all()}
        vt_order = [r.vehicle_type for r in VehicleCapacity.query.order_by(VehicleCapacity.sort_order).all()]
        vr_map  = {}
        for vr in VehicleRate.query.filter_by(center_code=main_code).all():
            vr_map.setdefault(vr.destination, {})[vr.vehicle_type] = vr.unit_price

        def _vehicle_cost(plt_dec, destination):
            rates = vr_map.get(destination or '', {})
            best  = None
            for vt in vt_order:
                if vt not in rates or vt not in caps:
                    continue
                mp   = caps[vt]
                up   = rates[vt]
                cost = up if plt_dec <= mp else up * math.ceil(plt_dec / mp)
                if best is None or cost < best:
                    best = cost
            return best or 0

        records = CalculationResult.query.filter_by(customer_id=customer_id).all()
        dest_agg = {}  # destination → {vt, cnt, plt, box, cost}
        for r in records:
            if (r.total_plt_decimal or 0) < threshold:
                continue
            cost = _vehicle_cost(r.total_plt_decimal or 0, r.destination)
            key  = (r.destination or '-', r.vehicle_type or '-')
            if key not in dest_agg:
                dest_agg[key] = {'cnt': 0, 'total_plt': 0.0, 'total_box': 0.0, 'total_cost': 0}
            dest_agg[key]['cnt']        += 1
            dest_agg[key]['total_plt']  += r.total_plt_decimal or 0
            dest_agg[key]['total_box']  += r.total_box_qty or 0
            dest_agg[key]['total_cost'] += cost

        result = sorted([
            {'destination': k[0], 'vehicle_type': k[1], 'cnt': v['cnt'],
             'total_plt': round(v['total_plt'], 2), 'total_box': round(v['total_box']),
             'total_cost': round(v['total_cost'])}
            for k, v in dest_agg.items()
        ], key=lambda x: -x['total_cost'])

    grand_total = sum(r['total_cost'] for r in result)
    return jsonify({'rows': result, 'grand_total': grand_total, 'threshold': threshold})


@app.route('/api/config', methods=['POST'])
def api_save_config():
    """시스템 설정 값을 저장 (지도 등 클라이언트에서 호출)"""
    data = request.get_json(silent=True) or {}
    key  = data.get('key', '').strip()
    val  = str(data.get('value', '')).strip()
    if not key or not val:
        return jsonify({'ok': False, 'error': 'key/value 필요'}), 400
    allowed = {'stops_per_vehicle', 'kakao_api_key', 'direct_plt_threshold'}
    if key not in allowed:
        return jsonify({'ok': False, 'error': '허용되지 않는 키'}), 403
    cfg = SystemConfig.query.filter_by(key=key).first()
    if cfg:
        cfg.value = val
    else:
        db.session.add(SystemConfig(key=key, value=val, description=''))
    db.session.commit()
    return jsonify({'ok': True, 'key': key, 'value': val})


@app.route('/api/vehicle-distance-rate')
def api_vehicle_distance_rate():
    """JS에서 루트별 비용 계산용 — 전체 단가표 반환"""
    from hub_rate_data import VEHICLE_TYPES, VEHICLE_PLT_CAP
    rows = VehicleDistanceRate.query.all()
    rates = {}
    for r in rows:
        rates.setdefault(r.vehicle_type, {})[r.km] = r.unit_price
    return jsonify({'rates': rates, 'plt_cap': VEHICLE_PLT_CAP, 'vehicle_types': VEHICLE_TYPES})


# ─── 시너지 분석 ──────────────────────────────────────────────────────────────

@app.route('/synergy')
def synergy_index():
    from sqlalchemy import func as sqlfunc
    customers = Customer.query.order_by(Customer.name).all()

    # 업로드된 배송지시서 현황
    total_own  = SynergyRoute.query.count()
    joint_own  = SynergyRoute.query.filter_by(car_flag=2).count()
    direct_own = SynergyRoute.query.filter_by(car_flag=1).count()

    # 고유 지역 수 (공동배송 기준)
    region_cnt = db.session.query(
        sqlfunc.count(sqlfunc.distinct(
            db.func.coalesce(SynergyRoute.sido + ' ' + SynergyRoute.sigungu, '')
        ))
    ).filter(SynergyRoute.car_flag == 2, SynergyRoute.sido.isnot(None)).scalar() or 0

    selected_id = request.args.get('customer_id', type=int)
    analysis = None
    selected_customer = None

    if selected_id and total_own > 0:
        selected_customer = Customer.query.get(selected_id)
        analysis = _run_synergy_analysis(selected_id)

    return render_template('synergy/index.html',
        customers=customers,
        total_own=total_own,
        joint_own=joint_own,
        direct_own=direct_own,
        region_cnt=region_cnt,
        selected_id=selected_id,
        selected_customer=selected_customer,
        analysis=analysis,
    )


def _run_synergy_analysis(customer_id):
    """자사 공동배송 루트 vs 화주사 배송지 시너지 분석"""
    # 자사 공동배송 루트: (sido, sigungu) → {plt, box, stores, region}
    own_joint = SynergyRoute.query.filter_by(car_flag=2).all()
    own_map = {}  # key: (sido, sigungu)
    for r in own_joint:
        if not r.sido or not r.sigungu:
            continue
        key = (r.sido, r.sigungu)
        if key not in own_map:
            own_map[key] = {
                'plt': 0.0, 'box': 0.0, 'stores': set(), 'regions': set()
            }
        own_map[key]['plt']   += r.plt_qty or 0
        own_map[key]['box']   += r.box_qty or 0
        own_map[key]['stores'].add(r.store_code or r.store_name or '')
        if r.delivery_region:
            own_map[key]['regions'].add(r.delivery_region.strip())

    # 화주사 배송지: 공동배송 건만 (직송은 시너지 분석 대상 아님)
    cust_results = CalculationResult.query.filter_by(
        customer_id=customer_id, delivery_mode='공동배송'
    ).all()

    match_map   = {}  # (sido, sigungu) → {cust_plt, cust_box, cust_cnt, own_info}
    unmatch_map = {}  # (sido, sigungu) → {cust_plt, cust_box, cust_cnt}

    for r in cust_results:
        sido, sigungu = extract_sido_sigungu(r.address or r.destination or '')
        if not sido or not sigungu:
            # destination 컬럼("경기도 수원시" 형태)으로 재시도
            if r.destination:
                parts = r.destination.strip().split()
                if len(parts) >= 2:
                    sido, sigungu = normalize_sido(parts[0]), parts[1]
        if not sido or not sigungu:
            continue
        sido = normalize_sido(sido)
        key  = (sido, sigungu)

        if key in own_map:
            if key not in match_map:
                match_map[key] = {
                    'sido': sido, 'sigungu': sigungu,
                    'cust_plt': 0.0, 'cust_box': 0.0, 'cust_cnt': 0,
                    'own_plt':  own_map[key]['plt'],
                    'own_box':  own_map[key]['box'],
                    'own_stores': len(own_map[key]['stores']),
                    'own_regions': ', '.join(sorted(own_map[key]['regions'])[:3]),
                }
            match_map[key]['cust_plt'] += r.total_plt_decimal or 0
            match_map[key]['cust_box'] += r.total_box_qty or 0
            match_map[key]['cust_cnt'] += 1
        else:
            if key not in unmatch_map:
                unmatch_map[key] = {
                    'sido': sido, 'sigungu': sigungu,
                    'cust_plt': 0.0, 'cust_box': 0.0, 'cust_cnt': 0,
                }
            unmatch_map[key]['cust_plt'] += r.total_plt_decimal or 0
            unmatch_map[key]['cust_box'] += r.total_box_qty or 0
            unmatch_map[key]['cust_cnt'] += 1

    total_cust = len(cust_results)
    match_cnt   = sum(v['cust_cnt'] for v in match_map.values())
    unmatch_cnt = sum(v['cust_cnt'] for v in unmatch_map.values())
    overlap_pct = round(match_cnt / total_cust * 100, 1) if total_cust else 0

    match_list   = sorted(match_map.values(),   key=lambda x: x['cust_plt'], reverse=True)
    unmatch_list = sorted(unmatch_map.values(), key=lambda x: x['cust_plt'], reverse=True)

    return {
        'total_cust':   total_cust,
        'match_cnt':    match_cnt,
        'unmatch_cnt':  unmatch_cnt,
        'overlap_pct':  overlap_pct,
        'match_regions':   len(match_map),
        'unmatch_regions': len(unmatch_map),
        'match_plt':    round(sum(v['cust_plt'] for v in match_map.values()), 1),
        'unmatch_plt':  round(sum(v['cust_plt'] for v in unmatch_map.values()), 1),
        'match_list':   match_list,
        'unmatch_list': unmatch_list,
    }


@app.route('/synergy/upload', methods=['POST'])
def synergy_upload():
    file = request.files.get('file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('synergy_index'))
    try:
        df = pd.read_excel(file)
        df.columns = [c.strip() for c in df.columns]

        overwrite = request.form.get('overwrite') == 'yes'
        if overwrite:
            SynergyRoute.query.delete()

        batch_id = str(uuid.uuid4())[:8]
        added = 0
        for _, row in df.iterrows():
            addr = str(row.get('ADDRESS', '') or '').strip()
            sido, sigungu = extract_sido_sigungu(addr)
            if sido:
                sido = normalize_sido(sido)

            def _date(col):
                v = row.get(col)
                if v is None or (isinstance(v, float) and math.isnan(v)):
                    return None
                try:
                    return pd.to_datetime(str(int(v)), format='%Y%m%d').date()
                except Exception:
                    return None

            def _flt(col):
                v = row.get(col)
                if v is None or (isinstance(v, float) and math.isnan(v)):
                    return None
                try:
                    return float(v)
                except Exception:
                    return None

            def _str(col):
                v = row.get(col)
                if v is None or (isinstance(v, float) and math.isnan(v)):
                    return None
                s = str(v).strip()
                return s if s and s.lower() != 'nan' else None

            db.session.add(SynergyRoute(
                batch_id        = batch_id,
                delivery_date   = _date('DELIVERY_DATE'),
                store_code      = _str('STORE_CODE'),
                store_name      = _str('STORE_NAME'),
                address         = addr or None,
                zipno           = _str('ZIPNO'),
                sido            = sido,
                sigungu         = sigungu,
                plt_qty         = _flt('PLT_QTY'),
                box_qty         = _flt('BOX_QTY'),
                car_flag        = int(_flt('CAR_FLAG') or 0),
                delivery_region = _str('DELIVERY_REGION_NAME'),
            ))
            added += 1

        db.session.commit()
        flash(f'배송지시서 {added:,}건 업로드 완료 (공동배송: {SynergyRoute.query.filter_by(car_flag=2).count():,}건)', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'업로드 오류: {e}', 'danger')
    return redirect(url_for('synergy_index'))


# ─── 공동배송 시뮬레이션 지도 ─────────────────────────────────────────────────

def _get_kakao_key():
    cfg = SystemConfig.query.filter_by(key='kakao_api_key').first()
    return cfg.value.strip() if cfg else ''


def _geocode_dest(destination):
    """배송지 시군구 → (lat, lon). 캐시 우선, 없으면 Kakao API 조회."""
    cached = DestinationCoord.query.get(destination)
    if cached:
        return cached.lat, cached.lon
    api_key = _get_kakao_key()
    if not api_key:
        return None, None
    try:
        resp = http_req.get(
            'https://dapi.kakao.com/v2/local/search/address.json',
            headers={'Authorization': f'KakaoAK {api_key}'},
            params={'query': destination, 'size': 1},
            timeout=4,
        )
        docs = resp.json().get('documents', [])
        if docs:
            lat, lon = float(docs[0]['y']), float(docs[0]['x'])
            db.session.add(DestinationCoord(destination=destination, lat=lat, lon=lon))
            db.session.commit()
            return lat, lon
    except Exception:
        pass
    return None, None


@app.route('/map')
def map_view():
    customer_id = request.args.get('customer_id', type=int)
    customer = Customer.query.get(customer_id) if customer_id else None
    centers = OurCenter.query.order_by(OurCenter.sort_order).all()
    stops_cfg = SystemConfig.query.get('stops_per_vehicle')
    stops_per_vehicle = int(stops_cfg.value) if stops_cfg else 8
    return render_template('map/simulation.html', centers=centers,
                           customer=customer, stops_per_vehicle=stops_per_vehicle)


@app.route('/api/map-data')
def api_map_data():
    customer_id = request.args.get('customer_id', type=int)
    threshold   = request.args.get('threshold', type=float)
    if customer_id:
        return _customer_map_data(customer_id, threshold=threshold)
    return _global_map_data()


def _global_map_data():
    centers_raw = OurCenter.query.order_by(OurCenter.sort_order).all()
    center_list = []
    for c in centers_raw:
        if c.lat and c.lon:
            center_list.append({
                'code':    c.center_code,
                'name':    c.center_name,
                'lat':     c.lat,
                'lon':     c.lon,
                'is_main': bool(c.is_main_center),
                'is_hub':  bool(c.is_direct_hub),
                'address': c.address or '',
            })
    center_map = {c['code']: c for c in center_list}

    transfer_routes = []
    seen_tr = set()
    for tr in TransferRate.query.all():
        key = tuple(sorted([tr.from_center_code, tr.to_center_code]))
        if key in seen_tr:
            continue
        seen_tr.add(key)
        fc = center_map.get(tr.from_center_code)
        tc = center_map.get(tr.to_center_code)
        if fc and tc:
            transfer_routes.append({
                'from_code': tr.from_center_code, 'to_code': tr.to_center_code,
                'from_name': fc['name'], 'to_name': tc['name'],
                'from_lat': fc['lat'], 'from_lon': fc['lon'],
                'to_lat': tc['lat'], 'to_lon': tc['lon'],
            })

    coord_cache = {dc.destination: (dc.lat, dc.lon) for dc in DestinationCoord.query.all()}
    center_dests = {}
    for vr in VehicleRate.query.all():
        dest = vr.destination
        lat, lon = coord_cache.get(dest, (None, None))
        entry = center_dests.setdefault(vr.center_code, {}).setdefault(dest, {
            'destination': dest, 'vehicles': set(), 'lat': lat, 'lon': lon
        })
        entry['vehicles'].add(vr.vehicle_type)

    delivery_routes = []
    for ccode, dests in center_dests.items():
        center = center_map.get(ccode)
        if not center:
            continue
        for dest, info in dests.items():
            if info['lat'] and info['lon']:
                delivery_routes.append({
                    'center_code': ccode, 'center_name': center['name'],
                    'destination': dest,
                    'lat': info['lat'], 'lon': info['lon'],
                    'vehicles': sorted(info['vehicles']),
                })

    all_dests = {vr.destination for vr in VehicleRate.query.all()}
    pending_geocode = [d for d in sorted(all_dests) if d not in coord_cache]

    return jsonify({
        'mode': 'global',
        'centers': center_list,
        'transfer_routes': transfer_routes,
        'delivery_routes': delivery_routes,
        'pending_geocode': pending_geocode,
        'stats': {
            'centers': len(center_list),
            'main_centers': sum(1 for c in center_list if c['is_main']),
            'transfer_routes': len(transfer_routes),
            'delivery_zones': len(all_dests),
            'geocoded': len(delivery_routes),
            'pending': len(pending_geocode),
        },
    })


def _haversine_km(lat1, lon1, lat2, lon2):
    """두 좌표 간 직선거리(km) — Haversine 공식"""
    R = 6371
    d = math.radians
    a = (math.sin(d(lat2 - lat1) / 2) ** 2
         + math.cos(d(lat1)) * math.cos(d(lat2)) * math.sin(d(lon2 - lon1) / 2) ** 2)
    return round(R * 2 * math.asin(math.sqrt(a)), 1)


def _customer_map_data(customer_id, _return_raw=False, threshold=None):
    from sqlalchemy import func as sqlfunc
    customer = Customer.query.get_or_404(customer_id)

    coord_cache = {dc.destination: (dc.lat, dc.lon) for dc in DestinationCoord.query.all()}

    # ── 출고내역 기반 배송지별 평균 PLT / BOX 집계 ────────────────────────────
    _sh_store_map = {}
    _sh_stores = StoreMaster.query.filter_by(customer_id=customer_id).count()
    _store_agg  = {}   # no-StoreMaster path에서 사용 (threshold 미지정 시)
    _cr_joint   = []   # threshold 지정 시 사용

    if threshold is not None:
        # ── threshold 지정: CalculationResult 기반 집계 (delivery_mode 무시, PLT < threshold) ──
        _cr_joint = (
            db.session.query(
                CalculationResult.destination,
                sqlfunc.sum(CalculationResult.total_plt_decimal).label('plt_sum'),
                sqlfunc.sum(CalculationResult.total_box_qty).label('box_sum'),
            )
            .filter(
                CalculationResult.customer_id == customer_id,
                CalculationResult.total_plt_decimal.isnot(None),
                CalculationResult.total_plt_decimal < threshold,
            )
            .group_by(CalculationResult.destination).all()
        )
        _total_biz_days = (
            db.session.query(sqlfunc.count(sqlfunc.distinct(CalculationResult.shipping_date)))
            .filter(CalculationResult.customer_id == customer_id).scalar() or 1
        )
        dest_plt_stats = {
            r.destination: {
                'avg_plt':  round(float(r.plt_sum or 0) / _total_biz_days, 3),
                'avg_box':  round(float(r.box_sum or 0) / _total_biz_days, 1),
                'ship_cnt': 0,
            }
            for r in _cr_joint if r.destination
        }

    elif _sh_stores:
        plt_rows = (
            db.session.query(
                StoreMaster.destination,
                sqlfunc.avg(ShippingHistory.plt_qty_decimal).label('avg_plt'),
                sqlfunc.avg(ShippingHistory.box_qty).label('avg_box'),
                sqlfunc.count(ShippingHistory.id).label('ship_cnt'),
            )
            .select_from(ShippingHistory)
            .join(
                StoreMaster,
                db.and_(
                    ShippingHistory.customer_id == StoreMaster.customer_id,
                    ShippingHistory.store_code  == StoreMaster.store_code,
                )
            )
            .filter(ShippingHistory.customer_id == customer_id)
            .group_by(StoreMaster.destination)
            .all()
        )
        dest_plt_stats = {
            r.destination: {
                'avg_plt':  round(float(r.avg_plt  or 0), 2),
                'avg_box':  round(float(r.avg_box  or 0), 1),
                'ship_cnt': int(r.ship_cnt or 0),
            }
            for r in plt_rows
        }
    else:
        # StoreMaster도 threshold도 없음 → store_code 단위 개별 핀 (점포별 좌표 조회)
        _vr_dests_raw = {vr.destination for vr in VehicleRate.query.with_entities(VehicleRate.destination).all()}

        _vr_norm_idx = {}
        for _d in _vr_dests_raw:
            _toks = _d.split()
            if _toks:
                _ns = _SIDO_NORM.get(_toks[0], _toks[0])
                _rest = ' '.join(_toks[1:])
                _vr_norm_idx[(_ns, _rest)] = _d

        def _addr_to_dest(addr):
            toks = (addr or '').strip().split()
            if len(toks) < 2:
                return None
            ns = _SIDO_NORM.get(toks[0], toks[0])
            for n in (2, 1):
                if len(toks) > n:
                    rest = ' '.join(toks[1:n + 1])
                    dest = _vr_norm_idx.get((ns, rest))
                    if dest:
                        return dest
            return None

        # 공동배송 점포만 사용 (직송 점포 제외)
        _joint_store_codes = {
            r.store_code for r in db.session.query(CalculationResult.store_code).filter(
                CalculationResult.customer_id == customer_id,
                CalculationResult.delivery_mode == '공동배송',
            ).distinct().all() if r.store_code
        }

        # 총 영업일수: 공동배송 기준 기대값 분모
        _total_biz_days = db.session.query(ShippingHistory.shipping_date).filter(
            ShippingHistory.customer_id == customer_id,
            ShippingHistory.store_code.in_(_joint_store_codes),
        ).distinct().count() or 1

        # store_code 단위 집계: 점포별 개별 지도 핀 (공동배송만)
        _store_agg = {}  # store_code → {box_sum, plt_sum, cnt, ship_dates, store_name, address, hub_dest}
        for sh in (db.session.query(ShippingHistory.address, ShippingHistory.box_qty,
                                    ShippingHistory.plt_qty_decimal, ShippingHistory.shipping_date,
                                    ShippingHistory.store_name, ShippingHistory.store_code)
                   .filter(ShippingHistory.customer_id == customer_id,
                           ShippingHistory.store_code.in_(_joint_store_codes)).all()):
            hub_dest = _addr_to_dest(sh.address or '')
            if not hub_dest:
                continue
            sc = sh.store_code or sh.address or 'unknown'
            if sc not in _store_agg:
                _store_agg[sc] = {
                    'box_sum': 0.0, 'plt_sum': 0.0, 'cnt': 0, 'ship_dates': set(),
                    'store_name': sh.store_name or sc,
                    'address': sh.address or '',
                    'hub_dest': hub_dest,
                }
            entry = _store_agg[sc]
            entry['cnt'] += 1
            if sh.shipping_date:
                entry['ship_dates'].add(sh.shipping_date)
            if sh.box_qty is not None:
                entry['box_sum'] += float(sh.box_qty)
            if sh.plt_qty_decimal is not None:
                entry['plt_sum'] += float(sh.plt_qty_decimal)

        # 기대값 = 총PLT ÷ 총영업일수: 하루 평균 기대 PLT/BOX
        dest_plt_stats = {
            sc: {
                'avg_plt':   round(v['plt_sum'] / _total_biz_days, 3),
                'avg_box':   round(v['box_sum'] / _total_biz_days, 1),
                'ship_days': len(v['ship_dates']),
                'ship_cnt':  v['cnt'],
            }
            for sc, v in _store_agg.items()
        }

    # ── 재고보관센터 ──────────────────────────────────────────────────────────
    sc_entry = CustomerStorageCenter.query.filter(
        CustomerStorageCenter.customer_name.ilike(f'%{customer.name}%')
    ).first()
    center_info = None
    storage_code = None
    if sc_entry:
        cobj = OurCenter.query.filter_by(center_code=sc_entry.center_code).first()
        if cobj and cobj.lat and cobj.lon:
            storage_code = sc_entry.center_code
            center_info = {
                'code': cobj.center_code, 'name': cobj.center_name,
                'lat': cobj.lat, 'lon': cobj.lon,
                'address': cobj.address or '',
                'sc_customer_name': sc_entry.customer_name,
            }

    # ── 이고 연결 센터 (TransferRate 기반) ────────────────────────────────────
    transfer_to_codes = set()
    if storage_code:
        for tr in TransferRate.query.filter_by(from_center_code=storage_code).all():
            transfer_to_codes.add(tr.to_center_code)

    # ── DeliveryZoneMapping 인메모리 캐시 ─────────────────────────────────────
    # VehicleRate destination은 "경기도 수원시" 같이 시군구 단위이므로
    # 읍면동 레벨 데이터를 시군구·시도 단위로 집계해 fallback 조회 지원
    zone_exact   = {}   # (sido, sigungu, dong) → center_code
    zone_sigungu = {}   # (sido, sigungu) → center_code  (첫 매칭 우선)
    zone_sido    = {}   # sido → center_code             (첫 매칭 우선)

    for m in DeliveryZoneMapping.query.all():
        s, sg, d, cc = m.sido, m.sigungu, m.eupmyeondong, m.center_code
        zone_exact[(s, sg, d)] = cc
        if (s, sg) not in zone_sigungu and sg:
            zone_sigungu[(s, sg)] = cc
        if s not in zone_sido:
            zone_sido[s] = cc

    has_zone_mapping = bool(zone_exact)

    def zone_lookup(sido, sigungu='', dong=''):
        sido    = (sido    or '').strip()
        sigungu = (sigungu or '').strip()
        dong    = (dong    or '').strip()
        # 1) 읍면동 정확 매칭
        if dong:
            v = zone_exact.get((sido, sigungu, dong))
            if v: return v
        # 2) 시군구 레벨 (빈 eupmyeondong 행)
        if sigungu:
            v = zone_exact.get((sido, sigungu, ''))
            if v: return v
        # 3) 시군구 집계 캐시 (읍면동 무관, 같은 시군구의 첫 데이터)
        if sigungu:
            v = zone_sigungu.get((sido, sigungu))
            if v: return v
        # 4) 시도 레벨
        return zone_sido.get(sido)

    # ── VehicleRate destination → center_code 역매핑 (fallback용) ────────────
    vr_dest_to_center = {}
    for vr in VehicleRate.query.with_entities(VehicleRate.destination, VehicleRate.center_code).all():
        if vr.destination not in vr_dest_to_center:
            vr_dest_to_center[vr.destination] = vr.center_code

    # ── destination string → (시도, 시군구) 분리 헬퍼 ────────────────────────
    def split_dest(dest):
        parts = (dest or '').strip().split()
        return (parts[0] if parts else ''), (parts[1] if len(parts) > 1 else '')

    # ── 배송 destination 수집 및 거점 센터 매핑 ───────────────────────────────
    # 우선순위:
    #   1. StoreMaster 점포가 있는 경우 → 점포의 destination만 사용
    #   2. 없는 경우 → VehicleRate 전체 destination 사용
    # 센터 결정 우선순위:
    #   A. DeliveryZoneMapping (지역별 담당 센터 명시)
    #   B. VehicleRate 역매핑 (차량단가에 등록된 센터)

    stores = StoreMaster.query.filter_by(customer_id=customer_id).all()

    # destination별 점포 목록 (팝업에 점포명·주소 표시용)
    dest_stores = {}
    for s in stores:
        if s.destination:
            dest_stores.setdefault(s.destination, []).append({
                'name':    s.store_name or '',
                'address': s.address    or '',
                'code':    s.store_code or '',
            })

    # zone_key → {hub_dest, geo_key, display} : store-level 핀에서 좌표 조회 및 센터 라우팅에 사용
    _zone_meta = {}

    if stores:
        # 점포 기반: 각 점포 destination을 센터에 매핑
        working = [(s.destination, s.sido or '', s.sigungu or '') for s in stores if s.destination]
    elif threshold is not None:
        # threshold 지정 + StoreMaster 없음: CalculationResult destination 기반
        working = [
            (r.destination, *split_dest(r.destination))
            for r in _cr_joint if r.destination
        ]
    else:
        # store_code 단위: 점포별 개별 핀
        working = [
            (sc, *split_dest(info['hub_dest']))
            for sc, info in _store_agg.items()
        ]
        dest_stores = {
            sc: [{'name': v['store_name'], 'address': v['address'], 'code': sc}]
            for sc, v in _store_agg.items()
        }
        _zone_meta = {
            sc: {
                'hub_dest': info['hub_dest'],
                'geo_key':  info['address'],
                'display':  info['store_name'],
            }
            for sc, info in _store_agg.items()
        }

    hub_zone_map = {}   # center_code → {zone_key → {destination, zone_key, geo_key, lat, lon}}
    for dest, sido, sigungu in working:
        if _SIDO_NORM.get(sido, sido) == '제주':
            continue
        # A. DeliveryZoneMapping
        center_code = zone_lookup(sido, sigungu) if has_zone_mapping else None
        # B. VehicleRate fallback (store-level 핀은 hub_dest로 조회)
        if not center_code:
            _fallback = _zone_meta.get(dest, {}).get('hub_dest', dest)
            center_code = vr_dest_to_center.get(_fallback)
        if not center_code:
            continue
        # 좌표: store-level은 전체 주소로, sigungu-level은 dest 문자열로 조회
        _geo_key = _zone_meta.get(dest, {}).get('geo_key', dest)
        lat, lon = coord_cache.get(_geo_key, (None, None))
        hub_zone_map.setdefault(center_code, {})[dest] = {
            'destination': _zone_meta.get(dest, {}).get('display', dest),
            'zone_key':    dest,
            'geo_key':     _geo_key,
            'lat': lat, 'lon': lon,
        }

    # ── OurCenter 코드 정규화 조회 캐시 ──────────────────────────────────────
    # DeliveryZoneMapping은 0002000 형식, OurCenter는 2000 형식 혼용
    # → exact match 후 int 변환 fallback으로 통합
    _center_lookup = {}
    for _c in OurCenter.query.all():
        _center_lookup[_c.center_code] = _c
        try:
            _int = int(_c.center_code)
            _center_lookup[str(_int)]        = _c   # "0002000" → "2000"
            _center_lookup[str(_int).zfill(7)] = _c # "2000"    → "0002000"
        except (ValueError, TypeError):
            pass

    def find_center(code):
        code = (code or '').strip()
        c = _center_lookup.get(code)
        if c: return c
        try:
            _i = int(code)
            return _center_lookup.get(str(_i)) or _center_lookup.get(str(_i).zfill(7))
        except (ValueError, TypeError):
            return None

    # ── hub_centers 구성 ─────────────────────────────────────────────────────
    hub_centers = []
    for center_code, dest_map in hub_zone_map.items():
        hobj = find_center(center_code)
        if not hobj or not hobj.lat or not hobj.lon:
            continue
        if hobj.center_code == storage_code:
            continue

        zones = []
        for v in dest_map.values():
            zk = v.get('zone_key', v['destination'])
            gk = v.get('geo_key',  v['destination'])
            if not v['lat']:
                continue
            stats = dest_plt_stats.get(zk, {})
            dist  = _haversine_km(hobj.lat, hobj.lon, v['lat'], v['lon'])
            zones.append({
                'destination': v['destination'],
                'lat':         v['lat'],
                'lon':         v['lon'],
                'distance_km': dist,
                'road_km':     round(dist * 1.3, 1),
                'avg_plt':     stats.get('avg_plt', 0),
                'avg_box':     stats.get('avg_box', 0),
                'ship_days':   stats.get('ship_days', 0),
                'ship_cnt':    stats.get('ship_cnt', 0),
                'stores':      dest_stores.get(zk, []),
            })
        pending_zones = []
        for pv in dest_map.values():
            if pv['lat']:
                continue
            pzk = pv.get('zone_key', pv['destination'])
            pst = dest_plt_stats.get(pzk, {})
            pending_zones.append({
                'geo':      pv.get('geo_key', pv['destination']),
                'name':     pv['destination'],
                'avg_plt':   pst.get('avg_plt', 0),
                'avg_box':   pst.get('avg_box', 0),
                'ship_days': pst.get('ship_days', 0),
                'ship_cnt':  pst.get('ship_cnt', 0),
                'stores':    dest_stores.get(pzk, []),
            })

        hub_centers.append({
            'code':          hobj.center_code,
            'name':          hobj.center_name,
            'lat':           hobj.lat,
            'lon':           hobj.lon,
            'address':       hobj.address or '',
            'zones':         zones,
            'pending_zones': pending_zones,
            'has_transfer':  hobj.center_code in transfer_to_codes,
            'is_main':       bool(hobj.is_main_center),
        })

    hub_centers.sort(key=lambda h: -len(h['zones']))

    if _return_raw:
        return hub_centers

    # ── 이고 전용 센터 (DeliveryZoneMapping/VehicleRate 담당지 없지만 이고 경유) ─
    hub_codes = {h['code'] for h in hub_centers}
    transfer_only_centers = []
    for code in transfer_to_codes:
        if code in hub_codes:
            continue
        tobj = find_center(code)
        if tobj and tobj.lat and tobj.lon:
            transfer_only_centers.append({
                'code': tobj.center_code, 'name': tobj.center_name,
                'lat': tobj.lat, 'lon': tobj.lon, 'address': tobj.address or '',
            })

    # 견적 통계 (패널 표시용)
    stat_rows = db.session.query(
        CalculationResult.delivery_mode,
        sqlfunc.count(CalculationResult.id).label('cnt'),
        sqlfunc.sum(CalculationResult.delivery_cost).label('cost'),
    ).filter(CalculationResult.customer_id == customer_id
    ).group_by(CalculationResult.delivery_mode).all()

    direct_cnt = sum(r.cnt for r in stat_rows if r.delivery_mode == '직송')
    joint_cnt  = sum(r.cnt for r in stat_rows if r.delivery_mode == '공동배송')
    total_cost = int(sum(r.cost or 0 for r in stat_rows))

    # pending_zones 항목이 {geo, name} 객체이거나 문자열 모두 지원
    _seen_geo = set()
    all_pending = []
    for hc in hub_centers:
        for z in hc['pending_zones']:
            geo = z['geo'] if isinstance(z, dict) else z
            if geo not in _seen_geo:
                _seen_geo.add(geo)
                all_pending.append(z)

    return jsonify({
        'mode': 'customer',
        'customer': {'id': customer_id, 'name': customer.name},
        'center': center_info,
        'hub_centers': hub_centers,
        'transfer_only_centers': transfer_only_centers,
        'data_source': 'stores' if stores else ('zone_mapping' if has_zone_mapping else 'vehicle_rate'),
        'pending_geocode': all_pending,
        'stats': {
            'hub_count': len(hub_centers),
            'transfer_hubs': len(transfer_only_centers),
            'store_count': len(stores),
            'direct': direct_cnt,
            'joint': joint_cnt,
            'total_cost': total_cost,
        },
    })


@app.route('/api/geocode-batch', methods=['POST'])
def api_geocode_batch():
    """미지오코딩 배송지 최대 N개 처리 후 결과 반환"""
    destinations = (request.json or {}).get('destinations', [])[:30]
    api_key = _get_kakao_key()
    results = []
    for dest in destinations:
        if DestinationCoord.query.get(dest):
            continue
        if not api_key:
            break
        try:
            resp = http_req.get(
                'https://dapi.kakao.com/v2/local/search/address.json',
                headers={'Authorization': f'KakaoAK {api_key}'},
                params={'query': dest, 'size': 1},
                timeout=4,
            )
            docs = resp.json().get('documents', [])
            if docs:
                lat, lon = float(docs[0]['y']), float(docs[0]['x'])
                db.session.add(DestinationCoord(destination=dest, lat=lat, lon=lon))
                results.append({'destination': dest, 'lat': lat, 'lon': lon})
        except Exception:
            pass
    if results:
        db.session.commit()
    return jsonify({'geocoded': results, 'count': len(results)})


# ─── 화주사별 재고보관센터 마스터 ─────────────────────────────────────────────

@app.route('/masters/customer-center')
def customer_center_master():
    mappings = CustomerStorageCenter.query.order_by(
        CustomerStorageCenter.customer_name, CustomerStorageCenter.center_name
    ).all()
    centers = OurCenter.query.order_by(OurCenter.sort_order).all()
    return render_template('masters/customer_center.html',
                           mappings=mappings, centers=centers)


@app.route('/masters/customer-center/add', methods=['POST'])
def customer_center_add():
    customer_code = request.form.get('customer_code', '').strip()
    customer_name = request.form.get('customer_name', '').strip()
    center_code   = request.form.get('center_code', '').strip()
    memo          = request.form.get('memo', '').strip() or None

    if not all([customer_code, customer_name, center_code]):
        flash('화주코드, 화주명, 센터는 필수 입력 항목입니다.', 'danger')
        return redirect(url_for('customer_center_master'))

    center = OurCenter.query.filter_by(center_code=center_code).first()
    if not center:
        flash('존재하지 않는 센터 코드입니다.', 'danger')
        return redirect(url_for('customer_center_master'))

    existing = CustomerStorageCenter.query.filter_by(
        customer_code=customer_code, center_code=center_code
    ).first()
    if existing:
        flash(f'[{customer_name}] — [{center.center_name}] 조합이 이미 등록되어 있습니다.', 'warning')
        return redirect(url_for('customer_center_master'))

    db.session.add(CustomerStorageCenter(
        customer_code=customer_code,
        customer_name=customer_name,
        center_code=center_code,
        center_name=center.center_name,
        memo=memo,
    ))
    db.session.commit()
    flash(f'[{customer_name}] → [{center.center_name}] 재고보관센터가 등록되었습니다.', 'success')
    return redirect(url_for('customer_center_master'))


@app.route('/masters/customer-center/<int:rid>/delete', methods=['POST'])
def customer_center_delete(rid):
    mapping = CustomerStorageCenter.query.get_or_404(rid)
    name = f'{mapping.customer_name} → {mapping.center_name}'
    db.session.delete(mapping)
    db.session.commit()
    flash(f'[{name}] 삭제되었습니다.', 'warning')
    return redirect(url_for('customer_center_master'))


@app.route('/masters/customer-center/<int:rid>/edit', methods=['POST'])
def customer_center_edit(rid):
    mapping = CustomerStorageCenter.query.get_or_404(rid)
    mapping.customer_code = request.form.get('customer_code', '').strip() or mapping.customer_code
    mapping.customer_name = request.form.get('customer_name', '').strip() or mapping.customer_name
    center_code = request.form.get('center_code', '').strip()
    if center_code:
        center = OurCenter.query.filter_by(center_code=center_code).first()
        if center:
            mapping.center_code = center_code
            mapping.center_name = center.center_name
    mapping.memo = request.form.get('memo', '').strip() or None
    db.session.commit()
    flash('수정되었습니다.', 'success')
    return redirect(url_for('customer_center_master'))


########################################################################
# 배송 거점 매핑 (공동배송 거점 센터 — 지역별)
########################################################################

import re as _re


def _parse_zone_line(line):
    """탭 혹은 다중 공백으로 구분된 '시도\t시군구\t읍면동\t센터명 (코드)' 한 줄을 파싱"""
    line = line.strip()
    if not line or '(' not in line:
        return None

    # 탭 우선, 없으면 2칸+ 공백으로 분리
    parts = line.split('\t')
    if len(parts) < 3:
        parts = _re.split(r'  +', line)

    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) < 2:
        return None

    # 마지막 파트에 (숫자코드) 포함 여부 확인
    m = _re.search(r'\((\d+)\)', parts[-1])
    if not m:
        return None

    center_code = m.group(1)
    center_name = _re.sub(r'\s*\(\d+\)\s*$', '', parts[-1]).strip()
    addr_parts  = parts[:-1]

    sido         = addr_parts[0] if len(addr_parts) > 0 else ''
    sigungu      = addr_parts[1] if len(addr_parts) > 1 else ''
    eupmyeondong = addr_parts[2] if len(addr_parts) > 2 else ''

    for bad in ('(비어 있음)', '(없음)', '-'):
        if sigungu      == bad: sigungu      = ''
        if eupmyeondong == bad: eupmyeondong = ''

    if not sido:
        return None

    return dict(sido=sido, sigungu=sigungu, eupmyeondong=eupmyeondong,
                center_code=center_code, center_name=center_name)


def lookup_delivery_center(sido, sigungu='', eupmyeondong=''):
    """주소(시도/시군구/읍면동) → 담당 거점센터 코드 반환 (없으면 None)
    읍면동 → 시군구 집계 → 시도 순으로 fallback
    """
    sido         = (sido         or '').strip()
    sigungu      = (sigungu      or '').strip()
    eupmyeondong = (eupmyeondong or '').strip()

    if eupmyeondong:
        r = DeliveryZoneMapping.query.filter_by(
            sido=sido, sigungu=sigungu, eupmyeondong=eupmyeondong).first()
        if r: return r.center_code

    if sigungu:
        # 시군구 명시 행 우선
        r = DeliveryZoneMapping.query.filter_by(
            sido=sido, sigungu=sigungu, eupmyeondong='').first()
        if r: return r.center_code
        # 같은 시군구의 아무 행이나 (읍면동 무관)
        r = DeliveryZoneMapping.query.filter_by(
            sido=sido, sigungu=sigungu).first()
        if r: return r.center_code

    # 시도 레벨
    r = DeliveryZoneMapping.query.filter_by(
        sido=sido, sigungu='', eupmyeondong='').first()
    if r: return r.center_code
    r = DeliveryZoneMapping.query.filter_by(sido=sido).first()
    return r.center_code if r else None


@app.route('/masters/delivery-zone-mapping', methods=['GET', 'POST'])
def delivery_zone_mapping():
    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'clear':
            cnt = DeliveryZoneMapping.query.count()
            DeliveryZoneMapping.query.delete()
            db.session.commit()
            flash(f'전체 {cnt:,}건 삭제 완료', 'warning')
            return redirect(url_for('delivery_zone_mapping'))

        if action == 'import':
            raw = ''
            f = request.files.get('tsv_file')
            if f and f.filename:
                raw = f.read().decode('utf-8-sig', errors='replace')
            else:
                raw = request.form.get('raw_data', '')

            if not raw.strip():
                flash('데이터가 없습니다.', 'danger')
                return redirect(url_for('delivery_zone_mapping'))

            # 기존 키 캐시
            existing = {
                (m.sido, m.sigungu, m.eupmyeondong)
                for m in DeliveryZoneMapping.query.with_entities(
                    DeliveryZoneMapping.sido,
                    DeliveryZoneMapping.sigungu,
                    DeliveryZoneMapping.eupmyeondong).all()
            }
            new_items, skipped, errors = [], 0, 0
            for line in raw.splitlines():
                row = _parse_zone_line(line)
                if not row:
                    if line.strip():
                        errors += 1
                    continue
                key = (row['sido'], row['sigungu'], row['eupmyeondong'])
                if key in existing:
                    skipped += 1
                else:
                    new_items.append(DeliveryZoneMapping(**row))
                    existing.add(key)

            if new_items:
                db.session.bulk_save_objects(new_items)
                db.session.commit()

            flash(
                f'{len(new_items):,}건 등록 완료'
                + (f' / {skipped:,}건 중복 스킵' if skipped else '')
                + (f' / {errors:,}건 파싱 오류' if errors else ''),
                'success' if new_items else 'info'
            )
            return redirect(url_for('delivery_zone_mapping'))

    stats = (
        db.session.query(
            DeliveryZoneMapping.center_name,
            DeliveryZoneMapping.center_code,
            db.func.count(DeliveryZoneMapping.id).label('cnt')
        )
        .group_by(DeliveryZoneMapping.center_code, DeliveryZoneMapping.center_name)
        .order_by(DeliveryZoneMapping.center_code)
        .all()
    )
    total   = DeliveryZoneMapping.query.count()
    centers = OurCenter.query.order_by(OurCenter.sort_order).all()
    return render_template('delivery_zone_mapping.html',
                           stats=stats, total=total, centers=centers)


@app.route('/api/zone-lookup')
def api_zone_lookup():
    sido    = request.args.get('sido', '').strip()
    sigungu = request.args.get('sigungu', '').strip()
    dong    = request.args.get('dong', '').strip()
    code    = lookup_delivery_center(sido, sigungu, dong)
    if not code:
        return jsonify({'center_code': None, 'center_name': None}), 404
    center = OurCenter.query.filter_by(center_code=code).first()
    return jsonify({
        'center_code': code,
        'center_name': center.center_name if center else code,
    })


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
